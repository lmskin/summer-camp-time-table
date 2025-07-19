#!/usr/bin/env python3
"""
Individual Student Timetable Generator

This script generates individual timetables for each student from a master Excel file
containing all students and room numbers. It uses the same logic as generate_student_timetables.py
but is designed to work with your existing Excel files.

Usage:
    python generate_individual_timetables.py

The script will automatically find and process all timetable files in the input directory
that match the pattern: {instrument}-{camp}-time-table.xlsx
"""

import re
import os
import datetime
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from shared_utils import sanitize_filename, process_sheet, load_student_name_mapping, load_room_no_mapping

# PDF conversion imports
try:
    import win32com.client as win32
    import pythoncom
    PDF_CONVERSION_AVAILABLE = True
except ImportError:
    print("Warning: PDF conversion not available. Install pywin32 to enable PDF export.")
    PDF_CONVERSION_AVAILABLE = False

def convert_excel_to_pdf(xlsx_file_path, pdf_file_path):
    """
    Convert Excel file to PDF using Excel COM automation.
    This preserves all formatting, merged cells, and styling.
    """
    if not PDF_CONVERSION_AVAILABLE:
        print(f"Skipping PDF conversion for {xlsx_file_path} - pywin32 not available")
        return False
        
    excel_app = None
    workbook = None
    
    try:
        # Initialize COM
        pythoncom.CoInitialize()
        
        # Convert paths to absolute paths
        abs_xlsx_path = os.path.abspath(xlsx_file_path)
        abs_pdf_path = os.path.abspath(pdf_file_path)
        
        # Create Excel application
        excel_app = win32.Dispatch("Excel.Application")
        excel_app.Visible = False
        excel_app.DisplayAlerts = False
        excel_app.ScreenUpdating = False
        
        # Open workbook
        workbook = excel_app.Workbooks.Open(abs_xlsx_path)
        
        # Get the active worksheet
        worksheet = workbook.ActiveSheet
        
        # Set page setup for landscape and fit to one page
        worksheet.PageSetup.Orientation = 2  # xlLandscape
        worksheet.PageSetup.Zoom = False
        worksheet.PageSetup.FitToPagesWide = 1
        worksheet.PageSetup.FitToPagesTall = 1
        worksheet.PageSetup.PrintArea = worksheet.UsedRange.Address
        
        # Set margins (in inches)
        worksheet.PageSetup.LeftMargin = excel_app.InchesToPoints(0.3)
        worksheet.PageSetup.RightMargin = excel_app.InchesToPoints(0.3)
        worksheet.PageSetup.TopMargin = excel_app.InchesToPoints(0.3)
        worksheet.PageSetup.BottomMargin = excel_app.InchesToPoints(0.3)
        worksheet.PageSetup.HeaderMargin = excel_app.InchesToPoints(0.1)
        worksheet.PageSetup.FooterMargin = excel_app.InchesToPoints(0.1)
        
        # Export to PDF
        workbook.ExportAsFixedFormat(0, abs_pdf_path)  # 0 = xlTypePDF
        
        # Close workbook and quit Excel
        workbook.Close(SaveChanges=False)
        excel_app.Quit()
        
        return True
        
    except Exception as e:
        print(f"Error converting {xlsx_file_path} to PDF: {e}")
        return False
        
    finally:
        # Clean up resources
        try:
            if workbook is not None:
                workbook.Close(SaveChanges=False)
        except:
            pass
        
        try:
            if excel_app is not None:
                excel_app.Quit()
        except:
            pass
        
        try:
            pythoncom.CoUninitialize()
        except:
            pass

def load_group_mappings(filename, music_instrument):
    """
    Loads group-to-student mappings from the specified CSV file.
    The CSV should have 'group_number' and 'student_no' columns.
    Since we're now working with student names instead of IDs, we'll try to map both.
    """
    group_mappings = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                group_number = row.get('group_number')
                student_nos_str = row.get('student_no')

                if group_number and student_nos_str:
                    group_name = f"Group {group_number.strip()}"
                    
                    # Extract student IDs based on instrument prefix (for backwards compatibility)
                    instrument_prefix = music_instrument[0].upper()
                    found_students = re.findall(rf'\b{instrument_prefix}\d+\b', student_nos_str)
                    
                    # Also try to extract student names (anything that's not an ID pattern)
                    # Split by common delimiters and clean up
                    potential_names = re.split(r'[,;&\n]+', student_nos_str)
                    for name in potential_names:
                        name = name.strip()
                        # If it's not a student ID pattern and has letters, consider it a name
                        if (name and 
                            not re.match(rf'^{instrument_prefix}\d+$', name) and
                            len(name) >= 2 and 
                            re.search(r'[A-Za-z]', name)):
                            found_students.append(name)
                    
                    if found_students:
                        if group_name not in group_mappings:
                            group_mappings[group_name] = []
                        group_mappings[group_name].extend(found_students)
    except FileNotFoundError:
        print(f"Warning: {filename} not found. No group activities will be added.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    return group_mappings

def load_room_mapping(filename):
    """
    Loads teacher-to-room mappings from the specified CSV file.
    The CSV should have 'teacher_name' and 'room_name' columns.
    """
    room_mappings = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                teacher_name = row.get('teacher_name')
                room_number = row.get('room_name')
                if teacher_name and room_number:
                    room_mappings[teacher_name.strip()] = room_number.strip()
    except FileNotFoundError:
        print(f"Warning: {filename} not found. No room numbers will be assigned to teachers.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    return room_mappings

def generate_individual_timetables(input_filename):
    """
    Main function to generate individual student timetables from a master Excel file.
    
    Args:
        input_filename (str): Path to the Excel file containing the master timetable
    """
    basename = os.path.basename(input_filename)
    print(f"\n=== Processing: {basename} ===")
    
    # Extract instrument from filename
    music_instrument = basename.split('-')[0].capitalize()
    print(f"Instrument: {music_instrument}")

    # Extract camp from filename
    camp_match = re.search(r"-(camp[ab])\-", basename, re.IGNORECASE)
    if not camp_match:
        print(f"Warning: Could not determine camp from filename {basename}. Cannot load mappings.")
        return
    
    camp_part = camp_match.group(1).lower()  # e.g., 'campa' or 'campb'
    print(f"Camp: {camp_part}")
    
    try:
        # Load the workbook
        workbook = load_workbook(input_filename, data_only=True)
        print(f"Loaded workbook with sheets: {workbook.sheetnames}")
    except FileNotFoundError:
        print(f"Error: {input_filename} not found.")
        return
    except Exception as e:
        print(f"Error loading workbook: {e}")
        return

    # Load mappings for the specific camp
    student_mapping_file = os.path.join("input", f"student_mapping-{camp_part}.csv")
    group_mapping_file = os.path.join("input", f"group_mapping-{camp_part}.csv")
    room_mapping_file = os.path.join("input", f"room_mapping-{camp_part}.csv")
    room_no_mapping_file = os.path.join("input", f"room_no_mapping-{camp_part}.csv")

    print(f"Loading mappings...")
    student_name_map = load_student_name_mapping(student_mapping_file)
    group_mappings = load_group_mappings(group_mapping_file, music_instrument)
    room_mappings = load_room_mapping(room_mapping_file)
    room_no_map = load_room_no_mapping(room_no_mapping_file)
    
    print(f"Loaded {len(student_name_map)} student names")
    print(f"Loaded {len(group_mappings)} groups")
    print(f"Loaded {len(room_mappings)} room mappings")
    print(f"Loaded {len(room_no_map)} room number mappings")
    
    # Create a reverse mapping from student to their groups
    student_to_groups = {}
    for group, students in group_mappings.items():
        for s in students:
            if s not in student_to_groups:
                student_to_groups[s] = set()
            student_to_groups[s].add(group)

    # Process all sheets and store their data
    processed_sheets = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        processed_sheets[sheet_name] = process_sheet(sheet)

    # Find all unique students across all processed sheets
    all_students = set()
    
    # Debug: Let's see what's actually in the cells
    print("🔍 Debugging cell contents...")
    sample_count = 0
    for sheet_name, sheet_data in processed_sheets.items():
        print(f"  Sheet: {sheet_name}")
        schedule_rows = sheet_data[2:]  # Schedule data starts from the third row
        for row_idx, row in enumerate(schedule_rows[:3]):  # Show first 3 rows as sample
            for col_idx, cell in enumerate(row[1:3]):  # Show first 2 columns as sample
                if cell and str(cell).strip():
                    print(f"    Row {row_idx+3}, Col {col_idx+2}: '{cell}'")
                    sample_count += 1
                    if sample_count >= 10:  # Limit debug output
                        break
            if sample_count >= 10:
                break
        if sample_count >= 10:
            break
    
    # Extract student names from cells (look for names that appear to be students)
    for sheet_name, sheet_data in processed_sheets.items():
        schedule_rows = sheet_data[2:]  # Schedule data starts from the third row
        for row in schedule_rows:
            for cell in row[1:]:
                if isinstance(cell, str) and cell.strip():
                    cell_content = cell.strip()
                    
                    # Skip obviously non-student entries
                    skip_patterns = [
                        r'^\d{1,2}:\d{2}',  # Time patterns
                        r'^(lunch|break|welcome|workshop|toilet|rehearsal|concert|briefing|yoga|regulation|masterclass|ensemble|practice|room|group)',  # Common activities
                        r'^(monday|tuesday|wednesday|thursday|friday|saturday|sunday)',  # Day names
                        r'^\d+$',  # Just numbers
                        r'^[A-Z]{1,3}\d+[A-Z]?$',  # Room codes like UG24, B123
                    ]
                    
                    should_skip = False
                    for pattern in skip_patterns:
                        if re.match(pattern, cell_content, re.IGNORECASE):
                            should_skip = True
                            break
                    
                    if should_skip:
                        continue
                    
                    # Look for what appears to be student names (contains letters, possibly spaces)
                    # Names should be at least 2 characters and contain letters
                    if (len(cell_content) >= 2 and 
                        re.search(r'[A-Za-z]', cell_content) and 
                        not re.search(r'^\d+$', cell_content)):
                        
                        # Extract individual names if multiple names are in one cell
                        # Split by common delimiters
                        potential_names = re.split(r'[,;&\n]+', cell_content)
                        
                        for name in potential_names:
                            name = name.strip()
                            if (len(name) >= 2 and 
                                re.search(r'[A-Za-z]', name) and
                                not re.match(r'^\d+$', name)):
                                all_students.add(name)

    print(f"Found {len(all_students)} students: {sorted(all_students)}")

    # Common activities that apply to all students
    common_activities = [
        "Welcome", "Lunch", "Break", "Ensemble Coaching", "Workshop", "Toilet Break",
        "Rehearsal for Students and Friends Concert", 
        "Lina Summer Camp of Music Students & Friends Concert",
        "After concert refreshment (Maritime Museum)", "Group Activity",
        "Briefing for Saturday", "Yoga Class", "Harp Regulation Workshop",
        "Harp Regulation Class", "Harp Regulation", "Cello Regulation & Maintenance Class",
        "Workshop - Warm Up", "Cello MasterClass", "MasterClass", "Flute MasterClass", "Harp MasterClass"
    ]

    # Create output directory
    output_dir = "student_timetables"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")

    # Define border style for cells
    thin_border = Border(
        left=Side(style='thin'), right=Side(style='thin'),
        top=Side(style='thin'), bottom=Side(style='thin')
    )

    # Determine start date and camp name based on filename
    start_date = None
    camp_name = ""
    if 'campa' in input_filename.lower():
        start_date = datetime.date(2025, 7, 14)
        camp_name = "CampA"
    elif 'campb' in input_filename.lower():
        start_date = datetime.date(2025, 7, 21)
        camp_name = "CampB"

    print(f"Processing timetables for {camp_name} starting {start_date}")

    # Generate timetable for each student
    for student_idx, student in enumerate(sorted(list(all_students)), 1):
        print(f"Processing student {student_idx}/{len(all_students)}: {student}")
        
        # Collect all time slots and daily schedules for this student
        all_time_slots = set()
        daily_schedules = {}

        # Process each day (sheet)
        for day_index, sheet_name in enumerate(workbook.sheetnames):
            is_day_6 = (day_index + 1 == 6)

            if sheet_name not in processed_sheets:
                continue
            
            sheet_data = processed_sheets[sheet_name]
            teachers = [str(name).strip() for name in sheet_data[0][1:]]
            schedule_rows = sheet_data[2:]
            
            student_schedule = []
            day_6_check_in_added = False

            for row in schedule_rows:
                time_val = row[0]
                time = ""
                
                # Handle different time formats from Excel
                if isinstance(time_val, datetime.time):
                    time = time_val.strftime('%H:%M')
                elif isinstance(time_val, datetime.datetime):
                    time = time_val.strftime('%H:%M')
                elif time_val is not None:
                    time_str = str(time_val).strip()
                    # Try to parse common time formats
                    for fmt in ['%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M:%S %p']:
                        try:
                            parsed_time = datetime.datetime.strptime(time_str, fmt).time()
                            time = parsed_time.strftime('%H:%M')
                            break
                        except ValueError:
                            continue
                    
                    # If no format worked, use the string as-is if it looks like a time
                    if not time and ':' in time_str:
                        time = time_str

                # Skip processing if time is empty or invalid
                if not time or time.strip() == '' or time.lower() in ['none', 'nan']:
                    continue

                # Skip times after 17:00 for all days
                try:
                    current_time_obj = datetime.datetime.strptime(time, '%H:%M').time()
                    if current_time_obj >= datetime.time(17, 0):
                        continue
                except ValueError:
                    pass

                activities = [str(act).strip() for act in row[1:]]
                activity_found_for_timeslot = False

                # Special handling for Day 6
                if is_day_6:
                    for activity in activities:
                        if not activity:
                            continue

                        # Special handling for "Check in" activity on Day 6
                        if "Check in Maritime Museum" in activity:
                            if not day_6_check_in_added:
                                check_in_activity = "Check in Maritime Museum\nBriefing for Saturday Concert\nMaritime Museum Tour"
                                student_schedule.extend([
                                    ("10:00", check_in_activity),
                                    ("10:15", check_in_activity),
                                    ("10:30", check_in_activity),
                                    ("10:45", check_in_activity)
                                ])
                                day_6_check_in_added = True
                            activity_found_for_timeslot = True
                            break
                        
                        # For all other activities on Day 6
                        student_schedule.append((time, activity))
                        activity_found_for_timeslot = True
                        break
                else:
                    # Regular day processing (Days 1-5)
                    for i, activity in enumerate(activities):
                        if not activity:
                            continue

                        # Handle Masterclass activities containing student's name
                        if not activity_found_for_timeslot and 'masterclass' in activity.lower() and student.lower() in activity.lower():
                            teacher = None
                            # Try to find teacher from the mapping in the activity string
                            for known_teacher in sorted(room_mappings.keys(), key=len, reverse=True):
                                if known_teacher in activity:
                                    teacher = known_teacher
                                    break
                            
                            # Fallback to header teacher
                            if not teacher:
                                teacher = teachers[i]

                            room_number = room_mappings.get(teacher, "TBD")

                            # Clean up activity description - remove student name
                            base_activity = activity
                            # Remove the student name (case insensitive)
                            base_activity = re.sub(re.escape(student), '', base_activity, flags=re.IGNORECASE).strip()
                            # Remove any existing room string
                            base_activity = re.sub(r'\s*\([^)]+\)$', '', base_activity).strip()
                            # Clean up extra commas and spaces
                            base_activity = re.sub(r'[,\s]+', ' ', base_activity).strip()

                            desc = f"{base_activity}\n({room_number})"
                            student_schedule.append((time, desc))
                            activity_found_for_timeslot = True

                        # Handle direct student name matches (private lessons, etc.)
                        if not activity_found_for_timeslot and student.lower() in activity.lower():
                            # Remove student name from activity
                            cleaned_activity = re.sub(re.escape(student), '', activity, flags=re.IGNORECASE).strip()
                            # Clean up extra commas and spaces
                            cleaned_activity = re.sub(r'^[,\s]+|[,\s]+$', '', cleaned_activity).strip()
                            
                            is_private_lesson = (activity.strip().lower() == student.lower()) or ('private lesson' in cleaned_activity.lower())
                            
                            # Check for "Lesson with {teacher} & pianist" pattern
                            pianist_lesson_match = re.search(r'lesson with (.+?) & pianist', cleaned_activity.lower())
                            if pianist_lesson_match:
                                original_match = re.search(r'lesson with (.+?) & pianist', activity, re.IGNORECASE)
                                teacher_name = original_match.group(1).strip() if original_match else pianist_lesson_match.group(1).strip()
                                column_teacher = teachers[i]
                                teacher_room = room_mappings.get(column_teacher, "TBD")
                                desc = f"Private Lesson with {teacher_name} & pianist\n({teacher_room})"
                                student_schedule.append((time, desc))
                                activity_found_for_timeslot = True
                            elif is_private_lesson:
                                teacher = teachers[i]
                                if teacher:
                                    # Find room for teacher
                                    room_number = ""
                                    for mapped_teacher, room in room_mappings.items():
                                        if mapped_teacher.lower() == teacher.lower():
                                            room_number = room
                                            break
                                    if not room_number:
                                        room_number = room_mappings.get(teacher, "")
                                    
                                    if not cleaned_activity:
                                        desc = f"Private Lesson with {teacher}"
                                    else:
                                        desc = cleaned_activity
                                    
                                    if room_number:
                                        desc += f"\n({room_number})"
                                else:
                                    desc = f"Practice\n({music_instrument} practice room)"
                                student_schedule.append((time, desc))
                            else:
                                if cleaned_activity.lower() == 'practice':
                                    cleaned_activity = f"Practice\n({music_instrument} practice room)"
                                student_schedule.append((time, cleaned_activity))
                            activity_found_for_timeslot = True

                        # Handle complex group activities
                        if not activity_found_for_timeslot and activity.lower().startswith('group') and "," in activity:
                            activity_body = activity[len('Group'):].strip()
                            parts = activity_body.replace(',', ' ').split()
                            
                            group_numbers = []
                            activity_name_parts = []
                            for part in parts:
                                if part.isdigit():
                                    group_numbers.append(part)
                                else:
                                    activity_name_parts.append(part)
                            
                            activity_name = ' '.join(activity_name_parts).strip()
                            involved_groups = {f"Group {num}" for num in group_numbers}
                            student_groups = student_to_groups.get(student, set())

                            if not student_groups.isdisjoint(involved_groups):
                                if 'acting class' in activity_name.lower():
                                    acting_room = room_mappings.get("Room Acting Class", "Room Acting Class")
                                    student_schedule.append((time, f"Acting Class\n({acting_room})"))
                                else:
                                    if 'group' in activity_name.lower() or 'room' in activity_name.lower():
                                        student_schedule.append((time, activity_name))
                                    else:
                                        student_schedule.append((time, f"{activity_name}\n(Group)"))
                                activity_found_for_timeslot = True

                        # Handle simple group activities
                        if not activity_found_for_timeslot and activity.lower().startswith('group'):
                            student_groups = student_to_groups.get(student, set())
                            
                            for group_name in student_groups:
                                if activity.startswith(group_name):
                                    room_match = re.search(r'\(Room\s+(.+?)\)', activity, re.IGNORECASE)
                                    if room_match:
                                        room_name = room_match.group(1)
                                        student_schedule.append((time, f"Ensemble\n(Room {room_name})"))
                                    else:
                                        teacher = teachers[i]
                                        room_number = "TBD"
                                        for mapped_teacher, room in room_mappings.items():
                                            if mapped_teacher.lower() == teacher.lower():
                                                room_number = room
                                                break
                                        if room_number == "TBD":
                                            room_number = room_mappings.get(teacher, "TBD")
                                        student_schedule.append((time, f"Ensemble\n({room_number})"))
                                    activity_found_for_timeslot = True
                                    break

                # Fallback for common activities
                if not activity_found_for_timeslot:
                    activity_to_add = None
                    for activity in activities:
                        if not activity:
                            continue
                        
                        # Skip MasterClass activities that contain other student names
                        if 'masterclass' in activity.lower():
                            # Check if this activity contains the current student's name
                            if student.lower() not in activity.lower():
                                continue  # Skip this MasterClass as it doesn't include current student
                        
                        for common_activity in common_activities:
                            if common_activity in activity:
                                activity_to_add = activity
                                break
                        if activity_to_add:
                            break
    
                    if activity_to_add:
                        student_schedule.append((time, activity_to_add))
                    else:
                        student_schedule.append((time, ""))

            # Sort schedule by time
            student_schedule.sort(key=lambda x: x[0])
            daily_schedules[sheet_name] = student_schedule
            for time, _ in student_schedule:
                all_time_slots.add(time)
        
        # Create the individual timetable Excel file
        sorted_times = sorted(list(all_time_slots))
        time_to_row = {time: i + 4 for i, time in enumerate(sorted_times)}

        student_wb = Workbook()
        student_ws = student_wb.active
        student_ws.title = "Full Timetable"

        # Add student name in row 1
        # If we have a mapping from student ID to name, try to use it
        # Otherwise, use the student name directly from the Excel file
        student_name = student  # Default to the name from Excel
        
        # Try to find a mapping if the student appears to be an ID
        if student in student_name_map:
            student_name = student_name_map[student]
        else:
            # If no direct mapping and student looks like a name already, use it
            # Also check if any mapping values match (reverse lookup)
            for student_id, mapped_name in student_name_map.items():
                if mapped_name.lower() == student.lower():
                    student_name = mapped_name
                    break
        
        student_ws.cell(row=1, column=1, value=student_name).font = Font(bold=True, size=28)
        
        # Add time column header
        student_ws.cell(row=3, column=1, value="Time").font = Font(bold=True, size=20)
        for i, time in enumerate(sorted_times):
            student_ws.cell(row=i + 4, column=1, value=time).font = Font(size=20)

        # Add daily schedules
        current_col = 2
        for day_index, sheet_name in enumerate(workbook.sheetnames):
            if sheet_name not in daily_schedules:
                continue
            
            # Create date header
            if start_date:
                current_date = start_date + datetime.timedelta(days=day_index)
                header_text = current_date.strftime('%d %B (%A)')
            else:
                header_text = sheet_name
            
            student_ws.cell(row=2, column=current_col, value=header_text).font = Font(bold=True, size=20)

            todays_schedule = daily_schedules[sheet_name]
            merged_cells_in_col = set()

            for idx, (time, activity) in enumerate(todays_schedule):
                if time not in time_to_row:
                    continue
                
                start_row = time_to_row[time]
                
                if start_row in merged_cells_in_col:
                    continue

                # Calculate row span for merging identical consecutive activities
                row_span = 1
                for next_time, next_activity in todays_schedule[idx+1:]:
                    if next_activity == activity:
                        row_span += 1
                    else:
                        break
                
                # Clean up activity text
                cell_activity = activity
                if activity == "DAY_6_FREE_TIME_BLOCK":
                    cell_activity = ""
                
                # Format room information on separate lines
                room_patterns = [
                    r'\s*(\(Room\s+[^)]+\))',
                    r'\s*(\([A-Z]{1,3}\d+[A-Z]?\))',
                    r'\s*(\([^)]*room[^)]*\))',
                    r'\s*(\(Group\))',
                    r'\s*(\([^)]*practice\s+room[^)]*\))'
                ]
                
                for pattern in room_patterns:
                    cell_activity = re.sub(pattern, r'\n\1', cell_activity, flags=re.IGNORECASE)
                
                # Handle "or" with proper line breaks
                cell_activity = re.sub(r'\s*\bor\b\s*', '\nor\n', cell_activity, flags=re.IGNORECASE)
                cell_activity = re.sub(r'\n+', '\n', cell_activity).strip()
                
                # Replace room names with room numbers
                if room_no_map:
                    for r_name, r_number in room_no_map.items():
                        cell_activity = cell_activity.replace(r_name, r_number)

                # Set cell value
                cell = student_ws.cell(row=start_row, column=current_col, value=cell_activity)

                # Merge cells for consecutive identical activities
                if row_span > 1:
                    end_row = start_row + row_span - 1
                    student_ws.merge_cells(start_row=start_row, start_column=current_col, end_row=end_row, end_column=current_col)
                    for r in range(start_row, end_row + 1):
                        merged_cells_in_col.add(r)

            current_col += 1

        # Merge student name across all columns in row 1
        student_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=current_col-1)

        # Set column widths and row heights
        for column in student_ws.columns:
            column_letter = get_column_letter(column[0].column)
            column_number = column[0].column
            
            if column_number == 1:  # Time column
                student_ws.column_dimensions[column_letter].width = 15
            else:  # Date columns
                student_ws.column_dimensions[column_letter].width = 80
            
        # Set row heights
        student_ws.row_dimensions[1].height = 50  # Student name header
        student_ws.row_dimensions[2].height = 30  # Date headers
        for row_index in range(3, student_ws.max_row + 1):
            student_ws.row_dimensions[row_index].height = 60  # Time and data rows

        # Apply borders, alignment, and fonts
        for row in student_ws.iter_rows(min_row=1, max_row=student_ws.max_row, min_col=1, max_col=student_ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                
                # Preserve student name font size
                if cell.row == 1 and cell.column == 1:
                    continue
                
                # Apply appropriate font size
                if cell.font and cell.font.bold:
                    cell.font = Font(bold=True, size=20)
                else:
                    cell.font = Font(size=20)

        # Save files
        sanitized_file_name = sanitize_filename(student_name)
        camp_part = f"_{camp_name}" if camp_name else ""
        
        # Save Excel file
        xlsx_file_path = os.path.join(output_dir, f'{sanitized_file_name}{camp_part}_timetable.xlsx')
        student_wb.save(xlsx_file_path)
        
        # Save PDF file
        pdf_file_path = os.path.join(output_dir, f'{sanitized_file_name}{camp_part}_timetable.pdf')
        convert_excel_to_pdf(xlsx_file_path, pdf_file_path)

    print(f"\n✅ Successfully generated timetables for {len(all_students)} students!")
    print(f"   📁 Output directory: {output_dir}")
    print(f"   📄 Files created: {len(all_students)} Excel files + {len(all_students)} PDF files")

def main():
    """
    Main function to find and process all timetable files.
    """
    print("🎼 Individual Student Timetable Generator")
    print("=" * 50)
    
    input_dir = "input"
    if not os.path.isdir(input_dir):
        print(f"❌ Error: Input directory '{input_dir}' not found!")
        return
    
    # Find all timetable files matching the expected pattern
    filename_pattern = re.compile(r"(cello|flute|harp)-(camp[ab])-time-table\.xlsx", re.IGNORECASE)
    timetable_files = [f for f in os.listdir(input_dir) if filename_pattern.match(f)]
    
    # Debug: Show all files in input directory
    print(f"📁 All files in '{input_dir}':")
    for f in os.listdir(input_dir):
        print(f"   • {f}")
        if filename_pattern.match(f):
            print(f"     ✅ Matches pattern")
        else:
            print(f"     ❌ Does not match pattern")

    if not timetable_files:
        print(f"❌ No timetable files found in '{input_dir}'")
        print("Expected pattern: {{instrument}}-{{camp}}-time-table.xlsx")
        print("Example: cello-campA-time-table.xlsx")
        return
    
    print(f"📁 Found {len(timetable_files)} timetable file(s):")
    for filename in sorted(timetable_files):
        print(f"   • {filename}")
    
    print("\n🔄 Starting processing...")
    
    for filename in sorted(timetable_files):
        full_path = os.path.join(input_dir, filename)
        try:
            generate_individual_timetables(full_path)
        except Exception as e:
            print(f"❌ Error processing {filename}: {e}")
            continue
    
    print("\n🎉 All files processed successfully!")

if __name__ == '__main__':
    main()
import re
import os
import itertools
import datetime
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from shared_utils import sanitize_filename, process_sheet, load_student_name_mapping, load_room_no_mapping

def load_room_mapping(filename):
    """
    Loads teacher-to-room mappings from the specified CSV file.
    The CSV should have 'teacher_name' and 'room_name' columns.
    """
    room_mappings = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                teacher_name = row.get('teacher_name')
                room_number = row.get('room_name')
                if teacher_name and room_number:
                    room_mappings[teacher_name.strip()] = room_number.strip()
    except FileNotFoundError:
        print(f"Warning: {filename} not found. No room numbers will be assigned to teachers.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    return room_mappings

def generate_teacher_timetables(input_filename):
    """
    Reads an Excel file with multiple sheets (each representing a date) and
    generates an individual Excel timetable for each teacher.
    """
    basename = os.path.basename(input_filename)
    music_instrument = basename.split('-')[0].capitalize()

    # Extract camp (e.g., "campA") from filename
    camp_match = re.search(r"-(camp[ab])\-", basename, re.IGNORECASE)
    if not camp_match:
        print(f"Warning: Could not determine camp from filename {basename}. Cannot load mappings.")
        return
    
    camp_part = camp_match.group(1) # e.g., 'campA' - preserve original case

    try:
        workbook = load_workbook(input_filename, data_only=True)
        print(f"\nProcessing teacher timetables for {basename}...")
    except FileNotFoundError:
        print(f"Error: {input_filename} not found.")
        return

    student_mapping_file = os.path.join("input", f"student_mapping-{camp_part}.csv")
    student_name_map = load_student_name_mapping(student_mapping_file)

    room_mapping_file = os.path.join("input", f"room_mapping-{camp_part}.csv")
    room_mappings = load_room_mapping(room_mapping_file)

    room_no_mapping_file = os.path.join("input", f"room_no_mapping-{camp_part}.csv")
    room_no_map = load_room_no_mapping(room_no_mapping_file)

    processed_sheets = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        processed_sheets[sheet_name] = process_sheet(sheet)

    all_teachers = set()
    for sheet_name, sheet_data in processed_sheets.items():
        if sheet_data:
            # Teacher names are in the first row, starting from the second column
            teacher_row = sheet_data[0]
            for teacher_name in teacher_row[1:]:
                if teacher_name and isinstance(teacher_name, str) and teacher_name.strip():
                    all_teachers.add(teacher_name.strip())

    output_dir = "teacher_timetables"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Determine start date and camp name based on filename
    start_date = None
    camp_name = ""
    if 'campa' in input_filename.lower():
        start_date = datetime.date(2025, 7, 14)
        camp_name = "CampA"
    elif 'campb' in input_filename.lower():
        start_date = datetime.date(2025, 7, 21)
        camp_name = "CampB"

    # Teachers with a special event on Friday evening
    special_friday_teachers = ["Stephane RETY", "Tomasz SKWERES", "Sivan MEGAN", "Liya HUANG", "Gwyneth WENTINK"]

    for teacher in sorted(list(all_teachers)):
        all_time_slots = set()
        daily_schedules = {}

        for day_index, sheet_name in enumerate(workbook.sheetnames):
            if sheet_name not in processed_sheets:
                continue
            
            is_day_6 = (day_index + 1) == 6
            sheet_data = processed_sheets[sheet_name]
            header = [str(h).strip() for h in sheet_data[0]]
            
            try:
                # Find the column index for the current teacher
                teacher_col_index = header.index(teacher)
            except ValueError:
                if is_day_6:
                    teacher_col_index = 1  # Assume the second column for Day 6 activities
                else:
                    # Teacher not present in this sheet
                    continue

            schedule_rows = sheet_data[2:]
            daily_schedule_map = {}
            
            for row in schedule_rows:
                time_val = row[0]
                time = time_val.strftime('%H:%M') if isinstance(time_val, datetime.time) else str(time_val).strip()
                
                if not time:
                    continue

                current_time_obj = None
                try:
                    current_time_obj = datetime.datetime.strptime(time, '%H:%M').time()
                    if current_time_obj >= datetime.time(22, 0):
                        continue

                    if is_day_6 and current_time_obj < datetime.time(11, 0):
                        continue
                except ValueError:
                    pass

                # Get the activity from the teacher's column
                activity = str(row[teacher_col_index]).strip() if len(row) > teacher_col_index else ""
                
                if is_day_6:
                    if "Lunch" in activity and "Dress Up, Warm Up" in activity:
                        activity = "Lunch"
                    elif "Concert call time" in activity:
                        activity = "Lunch"
                else: # Special processing only for days other than Day 6
                    if activity.lower().startswith("workshop") or activity.lower().startswith("briefing for saturday"):
                        activity = ""

                    if activity:
                        # Handle specific pattern: "{student_no} Private Lesson with {teacher name} & pianist"
                        instrument_prefix = music_instrument[0].upper()
                        private_lesson_pattern = rf'\b({instrument_prefix}\d+)\s+Private\s+Lesson\s+with\s+.+?\s+&\s+pianist'
                        match = re.search(private_lesson_pattern, activity, re.IGNORECASE)
                        if match:
                            student_id = match.group(1)
                            student_name = student_name_map.get(student_id, student_id)
                            activity = f"{student_name} with pianist"
                        else:
                            # Handle harp MasterClass activities without room numbers
                            if 'harp masterclass' in activity.lower() and 'by' in activity.lower():
                                # Extract teacher name from "Harp MasterClass by Teacher Name"
                                teacher_match = re.search(r'harp\s+masterclass\s+by\s+(.+?)(?:\*|$)', activity, re.IGNORECASE)
                                if teacher_match:
                                    masterclass_teacher = teacher_match.group(1).strip()
                                    # Remove any trailing asterisks or special characters
                                    masterclass_teacher = re.sub(r'\*+$', '', masterclass_teacher).strip()
                                    
                                    # Look up room for this teacher
                                    room_number = room_mappings.get(masterclass_teacher, "TBD")
                                    
                                    # Check if room info is already in the activity
                                    if '(' not in activity or ')' not in activity:
                                        # Add room information
                                        clean_activity = re.sub(r'\*+$', '', activity).strip()
                                        activity = f"{clean_activity}\n({room_number})"
                            
                            # Rename group activities like "Group 6" to "Group 6 Ensemble Coaching"
                            activity = re.sub(r'(Group\s+\d+)(?! Ensemble Coaching)', r'\1 Ensemble Coaching', activity)

                            # Find all student IDs (e.g., F1) in the activity string
                            student_ids = re.findall(rf'\b{instrument_prefix}\d+\b', activity)
                            
                            for student_id in student_ids:
                                # Replace each student ID with their name, if available
                                student_name = student_name_map.get(student_id, student_id)
                                activity = activity.replace(student_id, student_name)
                
                if activity:
                    daily_schedule_map[time] = activity

            # For weekdays, manually add the evening merge block
            if not is_day_6:
                is_friday = (day_index == 4)
                evening_times = [
                    "19:00", "19:15", "19:30", "19:45",
                    "20:00", "20:15", "20:30", "20:45",
                    "21:00", "21:15", "21:30", "21:45"
                ]
                
                evening_activity = "EVENING_MERGE_BLOCK"
                if is_friday and teacher in special_friday_teachers:
                    evening_activity = "Transfer to Mandarin Oriental"

                for evening_time in evening_times:
                    daily_schedule_map[evening_time] = evening_activity

            # For Saturday, manually add the morning merge block
            if is_day_6:
                morning_times = ["10:00", "10:15", "10:30", "10:45"]
                for morning_time in morning_times:
                    daily_schedule_map[morning_time] = "SATURDAY_MORNING_MERGE_BLOCK"

            # Sort the schedule by time to ensure correct grouping for merging
            teacher_schedule = sorted(daily_schedule_map.items())

            daily_schedules[sheet_name] = teacher_schedule
            for time, _ in teacher_schedule:
                all_time_slots.add(time)
        
        if not daily_schedules:
            continue

        sorted_times = sorted(list(all_time_slots))
        time_to_row = {time: i + 4 for i, time in enumerate(sorted_times)}

        teacher_wb = Workbook()
        teacher_ws = teacher_wb.active
        teacher_ws.title = "Full Timetable"

        # Add teacher name in row 1, merged across all columns
        teacher_ws.cell(row=1, column=1, value=teacher).font = Font(bold=True, size=14)
        
        teacher_ws.cell(row=3, column=1, value="Time").font = Font(bold=True, size=14)
        for i, time in enumerate(sorted_times):
            teacher_ws.cell(row=i + 4, column=1, value=time).font = Font(size=14)

        current_col = 2
        for day_index, sheet_name in enumerate(workbook.sheetnames):
            if sheet_name not in daily_schedules:
                continue
            
            if start_date:
                current_date = start_date + datetime.timedelta(days=day_index)
                header_text = current_date.strftime('%d %B (%A)')
            else:
                header_text = sheet_name

            teacher_ws.cell(row=2, column=current_col, value=header_text).font = Font(bold=True, size=14)
            
            # Create a full schedule for the day, including empty slots, to allow merging of consecutive empty cells.
            todays_schedule_map = dict(daily_schedules[sheet_name])
            full_day_schedule = [(time, todays_schedule_map.get(time, "")) for time in sorted_times]

            for activity, group in itertools.groupby(full_day_schedule, key=lambda x: x[1]):
                group_list = list(group)
                row_span = len(group_list)
                start_time = group_list[0][0]
                
                if start_time not in time_to_row: continue

                start_row = time_to_row[start_time]
                
                cell_activity = activity
                if activity == "EVENING_MERGE_BLOCK" or activity == "SATURDAY_MORNING_MERGE_BLOCK":
                    cell_activity = ""

                # Remove any "*" characters from the cell content
                cell_activity = cell_activity.replace('*', '')

                # Ensure room information is on a new line, but only if it's not already.
                if '(' in cell_activity and ')' in cell_activity and '\n(' not in cell_activity:
                    cell_activity = cell_activity.replace('(', '\n(', 1)

                # Ensure "Ensemble Coaching" is on a new line
                if 'Ensemble Coaching' in cell_activity and '\nEnsemble Coaching' not in cell_activity:
                    cell_activity = cell_activity.replace(' Ensemble Coaching', '\nEnsemble Coaching', 1)

                # Replace room names with room numbers
                if room_no_map:
                    for r_name, r_number in room_no_map.items():
                        cell_activity = cell_activity.replace(r_name, r_number)

                if row_span > 1:
                    end_row = start_row + row_span - 1
                    teacher_ws.merge_cells(start_row=start_row, start_column=current_col, end_row=end_row, end_column=current_col)

                # Set value on the top-left cell of the (merged) range
                cell = teacher_ws.cell(row=start_row, column=current_col, value=cell_activity)

            current_col += 1

        # Merge teacher name across all columns in row 1
        teacher_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=current_col-1)

        # Set column widths: Time column auto-fit, date columns set to 80
        for column in teacher_ws.columns:
            column_letter = get_column_letter(column[0].column)
            column_number = column[0].column
            
            if column_number == 1:  # Time column
                # Auto-fit time column based on content
                max_length = 0
                for cell in column:
                    try:
                        if cell.value:
                            lines = str(cell.value).split('\n')
                            max_line_length = max(len(line) for line in lines) if lines else 0
                            if max_line_length > max_length:
                                max_length = max_line_length
                    except:
                        pass
                
                # Set reasonable width for time column
                font_size_factor = 1.3
                padding = 2
                if max_length > 0:
                    adjusted_width = max(max_length * font_size_factor + padding, 15)
                    adjusted_width = min(adjusted_width, 25)  # Reasonable max for time column
                else:
                    adjusted_width = 15
                teacher_ws.column_dimensions[column_letter].width = adjusted_width
            else:  # Date columns (Monday to Saturday)
                # Set date columns to width 80
                teacher_ws.column_dimensions[column_letter].width = 80

        # Set specific row heights as requested
        teacher_ws.row_dimensions[1].height = 35  # Teacher name header
        teacher_ws.row_dimensions[2].height = 35  # Date headers
        for row_index in range(3, teacher_ws.max_row + 1):
            teacher_ws.row_dimensions[row_index].height = 35  # Time and data rows

        # Apply borders, alignment, and font to all cells
        for row in teacher_ws.iter_rows(min_row=1, max_row=teacher_ws.max_row, min_col=1, max_col=teacher_ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # Apply 14pt font to all cells, preserving existing bold formatting if any
                if cell.font and cell.font.bold:
                    cell.font = Font(bold=True, size=14)
                else:
                    cell.font = Font(size=14)

        sanitized_file_name = sanitize_filename(teacher)
        camp_part = f"_{camp_name}" if camp_name else ""
        file_path = os.path.join(output_dir, f'{sanitized_file_name}{camp_part}_timetable.xlsx')
        teacher_wb.save(file_path)

    print(f"Successfully generated timetables for {len(all_teachers)} teachers for {os.path.basename(input_filename)} in the '{output_dir}' directory.")

if __name__ == '__main__':
    input_dir = "input"
    if not os.path.isdir(input_dir):
        print(f"Error: Input directory '{input_dir}' not found or is not a directory.")
    else:
        # Regex to match the expected filename format for specific instruments
        filename_pattern = re.compile(r"(cello|flute|harp)-(camp[ab])\-time-table\.xlsx", re.IGNORECASE)
        
        timetable_files = [f for f in os.listdir(input_dir) if filename_pattern.match(f)]

        if not timetable_files:
            print(f"No timetable files matching the pattern '{{music-instrument}}-{{campA or campB}}-time-table.xlsx' were found in '{input_dir}'.")
        else:
            print(f"Found {len(timetable_files)} timetable file(s) to process: {', '.join(sorted(timetable_files))}")
            for filename in sorted(timetable_files):
                full_path = os.path.join(input_dir, filename)
                generate_teacher_timetables(full_path)

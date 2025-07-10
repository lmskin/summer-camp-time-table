import re
import os
import itertools
import datetime
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from shared_utils import sanitize_filename, process_sheet, load_student_name_mapping, load_room_no_mapping

def load_group_mappings(filename, music_instrument):
    """
    Loads group-to-student mappings from the specified CSV file.
    The CSV should have 'group_number' and 'student_no' columns.
    It extracts student IDs (e.g., F1) from the 'student_no' column.
    """
    group_mappings = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                group_number = row.get('group_number')
                student_nos_str = row.get('student_no')

                if group_number and student_nos_str:
                    group_name = f"Group {group_number.strip()}"
                    # Extract all F-numbers (e.g., F1, F23) from the string
                    instrument_prefix = music_instrument[0].upper()
                    found_students = re.findall(rf'\b{instrument_prefix}\d+\b', student_nos_str)
                    
                    if found_students:
                        if group_name not in group_mappings:
                            group_mappings[group_name] = []
                        group_mappings[group_name].extend(found_students)
    except FileNotFoundError:
        print(f"Warning: {filename} not found. No group activities will be added.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    return group_mappings

def load_room_mapping(filename):
    """
    Loads teacher-to-room mappings from the specified CSV file.
    The CSV should have 'teacher_name' and 'room_name' columns.
    """
    room_mappings = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                teacher_name = row.get('teacher_name')
                room_number = row.get('room_name')
                if teacher_name and room_number:
                    room_mappings[teacher_name.strip()] = room_number.strip()
    except FileNotFoundError:
        print(f"Warning: {filename} not found. No room numbers will be assigned to teachers.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    return room_mappings

def generate_timetables(input_filename):
    """
    Reads an Excel file with multiple sheets (each representing a date) and
    generates an individual Excel timetable for each student. Each student's
    timetable will contain multiple sheets, corresponding to the dates in the
    input file.
    """
    basename = os.path.basename(input_filename)
    music_instrument = basename.split('-')[0].capitalize()

    # Extract camp (e.g., "campA") from filename
    camp_match = re.search(r"-(camp[ab])\-", basename, re.IGNORECASE)
    if not camp_match:
        print(f"Warning: Could not determine camp from filename {basename}. Cannot load mappings.")
        return
    
    camp_part = camp_match.group(1) # e.g., 'campA' - preserve original case
    
    try:
        # Load the workbook once to process all sheets.
        # data_only=True ensures we get cell values instead of formulas.
        workbook = load_workbook(input_filename, data_only=True)
        print(f"\nProcessing student timetables for {basename}...")
    except FileNotFoundError:
        # This case is less likely now but good to keep as a safeguard
        print(f"Error: {input_filename} not found.")
        return

    # Load mappings for the specific camp
    student_mapping_file = os.path.join("input", f"student_mapping-{camp_part}.csv")
    group_mapping_file = os.path.join("input", f"group_mapping-{camp_part}.csv")
    room_mapping_file = os.path.join("input", f"room_mapping-{camp_part}.csv")

    student_name_map = load_student_name_mapping(student_mapping_file)
    group_mappings = load_group_mappings(group_mapping_file, music_instrument)
    room_mappings = load_room_mapping(room_mapping_file)
    
    room_no_mapping_file = os.path.join("input", f"room_no_mapping-{camp_part}.csv")
    room_no_map = load_room_no_mapping(room_no_mapping_file)
    
    # Create a reverse mapping from student to their groups for efficient lookup
    student_to_groups = {}
    for group, students in group_mappings.items():
        for s in students:
            if s not in student_to_groups:
                student_to_groups[s] = set()
            student_to_groups[s].add(group)

    # Process all sheets and store their data in a dictionary
    processed_sheets = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        processed_sheets[sheet_name] = process_sheet(sheet)

    # Find all unique students across all processed sheets
    all_students = set()
    for sheet_name, sheet_data in processed_sheets.items():
        schedule_rows = sheet_data[2:]  # Schedule data starts from the third row
        for row in schedule_rows:
            for cell in row[1:]:
                if isinstance(cell, str):
                    # Use regex to find all student IDs (e.g., F1) in a cell
                    instrument_prefix = music_instrument[0].upper()
                    found_students = re.findall(rf'\b{instrument_prefix}\d+\b', cell)
                    for s in found_students:
                        all_students.add(s)

    common_activities = [
        "Welcome",
        "Lunch",
        "Break",
        "Ensemble Coaching",
        "Workshop",
        "Toilet Break",
        "Rehearsal for Students and Friends Concert",
        "Lina Summer Camp of Music Students & Friends Concert",
        "After concert refreshment (Maritime Museum)",
        "Group Activity",
        "Briefing for Saturday",
        "Yoga Class",
        "Harp Regulation Workshop",
        "Harp Regulation Class",
        "Harp Regulation",
        "Cello Regulation & Maintance Class",
        "Cello Regulation & Maintenance Class",
        "Workshop - Warm Up",
        "Cello MasterClass",
        "MasterClass",
        "Flute MasterClass",
        "Harp MasterClass"
    ]

    output_dir = "student_timetables"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    # Determine start date and camp name based on filename
    start_date = None
    camp_name = ""
    if 'campa' in input_filename.lower():
        start_date = datetime.date(2025, 7, 14)
        camp_name = "CampA"
    elif 'campb' in input_filename.lower():
        start_date = datetime.date(2025, 7, 21)
        camp_name = "CampB"

    # Generate one single-sheet Excel file for each student
    for student in sorted(list(all_students)):
        # Pre-process to gather all time slots and daily schedules for the student
        all_time_slots = set()
        daily_schedules = {}

        # Using workbook.sheetnames to preserve the order of days
        for day_index, sheet_name in enumerate(workbook.sheetnames):
            is_day_6 = (day_index + 1 == 6)

            if sheet_name not in processed_sheets:
                continue
            
            sheet_data = processed_sheets[sheet_name]
            teachers = [str(name).strip() for name in sheet_data[0][1:]]
            schedule_rows = sheet_data[2:]
            
            student_schedule = []
            day_6_check_in_added = False  # Flag to ensure it's added only once
            
            is_friday = (day_index == 4)
            is_harp = music_instrument.lower() == 'harp'

            for row in schedule_rows:
                time_val = row[0]
                time = ""
                
                # Handle different time formats from Excel
                if isinstance(time_val, datetime.time):
                    time = time_val.strftime('%H:%M')
                elif isinstance(time_val, datetime.datetime):
                    time = time_val.strftime('%H:%M')
                elif time_val is not None:
                    time_str = str(time_val).strip()
                    # Try to parse common time formats
                    for fmt in ['%H:%M:%S', '%H:%M', '%I:%M %p', '%I:%M:%S %p']:
                        try:
                            parsed_time = datetime.datetime.strptime(time_str, fmt).time()
                            time = parsed_time.strftime('%H:%M')
                            break
                        except ValueError:
                            continue
                    
                    # If no format worked, use the string as-is if it looks like a time
                    if not time and ':' in time_str:
                        time = time_str

                # Skip processing if time is empty or invalid
                if not time or time.strip() == '' or time.lower() in ['none', 'nan']:
                    continue

                # For Day 1-5, students finish at 17:00. For Day 6, they finish at 17:00.
                if day_index < 5:  # Day 1 to 5
                    try:
                        if datetime.datetime.strptime(time, '%H:%M').time() >= datetime.time(17, 0):
                            continue  # Skip this timeslot
                    except ValueError:
                        pass  # Not a time format
                elif is_day_6:  # Day 6
                    try:
                        current_time_obj = datetime.datetime.strptime(time, '%H:%M').time()
                        if current_time_obj >= datetime.time(17, 0):
                            continue  # Skip this timeslot
                        
                        # Handle the 16:30-17:00 merge block
                        if datetime.time(16, 30) <= current_time_obj < datetime.time(17, 0):
                            student_schedule.append((time, "DAY_6_FREE_TIME_BLOCK"))
                            continue # Skip other processing for this row
                    except ValueError:
                        pass  # Not a time format

                activities = [str(act).strip() for act in row[1:]]
                activity_found_for_timeslot = False

                if is_day_6:
                    # For Day 6, any activity is considered a common activity for all students.
                    # Find the first non-empty activity in the row.
                    for activity in activities:
                        if not activity:
                            continue

                        # Special handling for "Check in" activity
                        if "Check in Maritime Museum" in activity:
                            if not day_6_check_in_added:
                                check_in_activity = "Check in Maritime Museum\nBriefing for Saturday Concert\nMaritime Museum Tour"
                                student_schedule.extend([
                                    ("10:00", check_in_activity),
                                    ("10:15", check_in_activity),
                                    ("10:30", check_in_activity),
                                    ("10:45", check_in_activity)
                                ])
                                day_6_check_in_added = True
                            # Once the block is added, we don't need to process this specific activity again.
                            # We break here to process the next time slot from the source file.
                            activity_found_for_timeslot = True
                            break
                        
                        # For all other activities on Day 6
                        student_schedule.append((time, activity))
                        activity_found_for_timeslot = True
                        break # Found an activity for this time slot, move to the next.
                else:
                    # For all other days, run the specific matching logic.
                    for i, activity in enumerate(activities):
                        if not activity:
                            continue

                        # Generalized logic for any Masterclass containing student's ID
                        if not activity_found_for_timeslot and 'masterclass' in activity.lower() and re.search(r'\b' + re.escape(student) + r'\b', activity):
                            teacher = None
                            # Try to find a teacher from the mapping directly in the activity string
                            # Iterate a sorted list of teacher names (longest first) to avoid substring conflicts
                            for known_teacher in sorted(room_mappings.keys(), key=len, reverse=True):
                                if known_teacher in activity:
                                    teacher = known_teacher
                                    break
                            
                            # Fallback to header if no teacher found in string (less reliable)
                            if not teacher:
                                teacher = teachers[i]

                            room_number = room_mappings.get(teacher, "TBD")

                            # Remove all student IDs (e.g., H1, F12) from the activity string
                            instrument_prefix = music_instrument[0].upper()
                            base_activity = re.sub(rf'\b{instrument_prefix}\d+\b,?\s*', '', activity).strip()
                            
                            # Also remove any existing room string, since we'll add the correct one from the mapping
                            base_activity = re.sub(r'\s*\([^)]+\)$', '', base_activity).strip()

                            desc = f"{base_activity}\n({room_number})"
                            
                            student_schedule.append((time, desc))
                            activity_found_for_timeslot = True

                        # Priority 1: Direct student match (will be skipped if the above logic runs)
                        if not activity_found_for_timeslot and re.search(r'\b' + re.escape(student) + r'\b', activity):
                            cleaned_activity = re.sub(r'\b' + re.escape(student) + r'\b', '', activity).strip()
                            is_private_lesson = (activity.strip() == student) or ('private lesson' in cleaned_activity.lower())
                            
                            # Check for "Lesson with {teacher} & pianist" pattern
                            pianist_lesson_match = re.search(r'lesson with (.+?) & pianist', cleaned_activity.lower())
                            if pianist_lesson_match:
                                # Extract teacher name from the original activity (not cleaned_activity) to preserve exact formatting
                                original_match = re.search(r'lesson with (.+?) & pianist', activity, re.IGNORECASE)
                                if original_match:
                                    teacher_name = original_match.group(1).strip()
                                else:
                                    teacher_name = pianist_lesson_match.group(1).strip()
                                # Use the column header teacher's room instead of the teacher mentioned in the activity
                                column_teacher = teachers[i]
                                teacher_room = room_mappings.get(column_teacher, "TBD")
                                
                                desc = f"Private Lesson with {teacher_name} & pianist\n({teacher_room})"
                                student_schedule.append((time, desc))
                                activity_found_for_timeslot = True
                            elif is_private_lesson:
                                teacher = teachers[i]
                                if teacher:
                                    # Case-insensitive room mapping lookup
                                    room_number = ""
                                    for mapped_teacher, room in room_mappings.items():
                                        if mapped_teacher.lower() == teacher.lower():
                                            room_number = room
                                            break
                                    if not room_number:
                                        room_number = room_mappings.get(teacher, "")
                                    
                                    # If the activity was just the student ID, create a default description.
                                    # Otherwise, use the cleaned activity text which might contain more details.
                                    if not cleaned_activity:
                                        desc = f"Private Lesson with {teacher}"
                                    else:
                                        desc = cleaned_activity
                                    
                                    if room_number:
                                        desc += f"\n({room_number})"
                                else:
                                    # Fallback if no teacher is specified in the column for a private lesson
                                    desc = f"Practice\n({music_instrument} practice room)"
                                student_schedule.append((time, desc))
                            else:
                                # It's some other activity involving the student (e.g., a duet or practice)
                                if cleaned_activity.lower() == 'practice':
                                    cleaned_activity = f"Practice\n({music_instrument} practice room)"
                                student_schedule.append((time, cleaned_activity))
                            activity_found_for_timeslot = True

                        # Priority 2: Complex group match
                        if not activity_found_for_timeslot and activity.lower().startswith('group') and "," in activity:
                            activity_body = activity[len('Group'):].strip()
                            
                            parts = activity_body.replace(',', ' ').split()
                            
                            group_numbers = []
                            activity_name_parts = []
                            for part in parts:
                                if part.isdigit():
                                    group_numbers.append(part)
                                else:
                                    activity_name_parts.append(part)
                            
                            activity_name = ' '.join(activity_name_parts).strip()
                            involved_groups = {f"Group {num}" for num in group_numbers}

                            student_groups = student_to_groups.get(student, set())

                            if not student_groups.isdisjoint(involved_groups):
                                if 'acting class' in activity_name.lower():
                                    # Use the room mapping to find the correct room for acting class
                                    acting_room = room_mappings.get("Room Acting Class", "Room Acting Class")
                                    student_schedule.append((time, f"Acting Class\n({acting_room})"))
                                else:
                                    # If the activity name already implies it's a group or has a room, don't add "(Group)"
                                    if 'group' in activity_name.lower() or 'room' in activity_name.lower():
                                        student_schedule.append((time, activity_name))
                                    else:
                                        student_schedule.append((time, f"{activity_name}\n(Group)"))
                                activity_found_for_timeslot = True

                        # Priority 3: Simple group match (e.g., "Group 1")
                        if not activity_found_for_timeslot and activity.lower().startswith('group'):
                            student_groups = student_to_groups.get(student, set())
                            
                            # New logic for group activities
                            for group_name in student_groups:
                                if activity.startswith(group_name):
                                    room_match = re.search(r'\(Room\s+(.+?)\)', activity, re.IGNORECASE)
                                    if room_match:
                                        room_name = room_match.group(1)
                                        student_schedule.append((time, f"Ensemble\n(Room {room_name})"))
                                    else:
                                        teacher = teachers[i]
                                        # Case-insensitive room mapping lookup
                                        room_number = "TBD"
                                        for mapped_teacher, room in room_mappings.items():
                                            if mapped_teacher.lower() == teacher.lower():
                                                room_number = room
                                                break
                                        if room_number == "TBD":
                                            room_number = room_mappings.get(teacher, "TBD")
                                        student_schedule.append((time, f"Ensemble\n({room_number})"))
                                    activity_found_for_timeslot = True
                                    break  # Found a match, no need to check other groups
                
                # Fallback for common activities or Free Time
                if not activity_found_for_timeslot:
                    # Check for any common activity for this timeslot.
                    activity_to_add = None
                    for activity in activities:
                        if not activity:
                            continue
                        
                        # Skip MasterClass activities that contain specific student IDs but don't include current student
                        if 'masterclass' in activity.lower():
                            # Check if this activity contains any student IDs
                            instrument_prefix = music_instrument[0].upper()
                            found_students = re.findall(rf'\b{instrument_prefix}\d+\b', activity)
                            if found_students and student not in found_students:
                                continue  # Skip this MasterClass as it doesn't include current student
                        
                        for common_activity in common_activities:
                            if common_activity in activity:
                                activity_to_add = activity
                                break  # Found a common activity
                        if activity_to_add:
                            break
    
                    if activity_to_add:
                        student_schedule.append((time, activity_to_add))
                    else:
                        # Debug output for student C1 - no activity found at all
                        student_schedule.append((time, ""))

            # Sort the schedule by time to ensure correct grouping for merging
            student_schedule.sort(key=lambda x: x[0])

            daily_schedules[sheet_name] = student_schedule
            for time, _ in student_schedule:
                all_time_slots.add(time)
        
        sorted_times = sorted(list(all_time_slots))
        time_to_row = {time: i + 3 for i, time in enumerate(sorted_times)}

        student_wb = Workbook()
        student_ws = student_wb.active
        student_ws.title = "Full Timetable"

        student_ws.cell(row=2, column=1, value="Time").font = Font(bold=True, size=14)
        for i, time in enumerate(sorted_times):
            student_ws.cell(row=i + 3, column=1, value=time).font = Font(size=14)

        current_col = 2
        for day_index, sheet_name in enumerate(workbook.sheetnames):
            if sheet_name not in daily_schedules:
                continue
            
            if start_date:
                current_date = start_date + datetime.timedelta(days=day_index)
                header_text = current_date.strftime('%d %B (%A)')
            else:
                header_text = sheet_name
            
            student_ws.cell(row=1, column=current_col, value=header_text).font = Font(bold=True, size=14)

            todays_schedule = daily_schedules[sheet_name]
            
            # Keep track of merged cells to avoid writing to them again
            merged_cells_in_col = set()

            for idx, (time, activity) in enumerate(todays_schedule):
                if time not in time_to_row:
                    continue
                
                start_row = time_to_row[time]
                
                if start_row in merged_cells_in_col:
                    continue

                # Calculate row_span for merging
                row_span = 1
                for next_time, next_activity in todays_schedule[idx+1:]:
                    if next_activity == activity:
                        row_span += 1
                    else:
                        break
                
                cell_activity = activity
                if activity == "DAY_6_FREE_TIME_BLOCK":
                    cell_activity = ""
                
                # Ensure room information is always on a separate line
                # Look for room patterns and move them to new lines if they're not already
                room_patterns = [
                    r'\s*(\(Room\s+[^)]+\))',  # (Room 246), (Room UG24), etc.
                    r'\s*(\([A-Z]{1,3}\d+[A-Z]?\))',  # (UG24), (LG1), (B123), etc.
                    r'\s*(\([^)]*room[^)]*\))',  # Any parentheses containing "room"
                    r'\s*(\(Group\))',  # (Group)
                    r'\s*(\([^)]*practice\s+room[^)]*\))',  # Practice room references
                    r'\s+(or)\s+'  # The word "or" surrounded by spaces
                ]
                
                for pattern in room_patterns:
                    # Replace inline room info with newline + room info
                    cell_activity = re.sub(pattern, r'\n\1', cell_activity, flags=re.IGNORECASE)
                
                # Clean up any double newlines or leading/trailing whitespace
                cell_activity = re.sub(r'\n+', '\n', cell_activity).strip()
                
                # Replace room names with room numbers
                if room_no_map:
                    for r_name, r_number in room_no_map.items():
                        cell_activity = cell_activity.replace(r_name, r_number)

                cell = student_ws.cell(row=start_row, column=current_col, value=cell_activity)

                if row_span > 1:
                    end_row = start_row + row_span - 1
                    student_ws.merge_cells(start_row=start_row, start_column=current_col, end_row=end_row, end_column=current_col)
                    # Mark cells as merged
                    for r in range(start_row, end_row + 1):
                        merged_cells_in_col.add(r)

            current_col += 1

        # Auto-fit column widths based on content
        for column in student_ws.columns:
            column_letter = get_column_letter(column[0].column)
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        # Count lines and find the longest line
                        lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in lines) if lines else 0
                        if max_line_length > max_length:
                            max_length = max_line_length
                except:
                    pass
            
            # Set a reasonable width with some padding
            adjusted_width = min(max(max_length + 2, 15), 60)  # Min 15, max 60 characters with padding
            student_ws.column_dimensions[column_letter].width = adjusted_width
            
        # Calculate uniform row height based on maximum content across entire worksheet
        max_height_needed = 15  # Minimum row height
        for row_index in range(1, student_ws.max_row + 1):
            for col_index in range(1, student_ws.max_column + 1):
                cell = student_ws.cell(row=row_index, column=col_index)
                if cell.value:
                    # Count lines from explicit newlines
                    lines = str(cell.value).count('\n') + 1
                    
                    # Estimate height based on line count (15 pixels per line as a heuristic)
                    estimated_height = lines * 15

                    if estimated_height > max_height_needed:
                        max_height_needed = estimated_height
        
        # Apply uniform height to all rows
        for row_index in range(1, student_ws.max_row + 1):
            student_ws.row_dimensions[row_index].height = max_height_needed


        # Apply borders, alignment, and font to all cells
        for row in student_ws.iter_rows(min_row=1, max_row=student_ws.max_row, min_col=1, max_col=student_ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # Apply 14pt font to all cells, preserving existing bold formatting if any
                if cell.font and cell.font.bold:
                    cell.font = Font(bold=True, size=14)
                else:
                    cell.font = Font(size=14)

        # Use student name for the filename, falling back to student number
        student_name = student_name_map.get(student, student)
        sanitized_file_name = sanitize_filename(student_name)
        camp_part = f"_{camp_name}" if camp_name else ""
        file_path = os.path.join(output_dir, f'{sanitized_file_name}{camp_part}_timetable.xlsx')
        student_wb.save(file_path)

    print(f"Successfully generated timetables for {len(all_students)} students for {os.path.basename(input_filename)} in the '{output_dir}' directory.")


if __name__ == '__main__':
    input_dir = "input"
    if not os.path.isdir(input_dir):
        print(f"Error: Input directory '{input_dir}' not found or is not a directory.")
    else:
        # Regex to match the expected filename format for specific instruments
        filename_pattern = re.compile(r"(cello|flute|harp)-(camp[ab])\-time-table\.xlsx", re.IGNORECASE)
        
        timetable_files = [f for f in os.listdir(input_dir) if filename_pattern.match(f)]

        if not timetable_files:
            print(f"No timetable files matching the pattern '{{music-instrument}}-{{campA or campB}}-time-table.xlsx' were found in '{input_dir}'.")
        else:
            print(f"Found {len(timetable_files)} timetable file(s) to process: {', '.join(sorted(timetable_files))}")
            for filename in sorted(timetable_files):
                full_path = os.path.join(input_dir, filename)
                generate_timetables(full_path) 
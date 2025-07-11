import os
import re
import csv
import shutil
from openpyxl import load_workbook
from openpyxl.styles import Alignment
from copy import copy

# Assuming shared_utils.py is in the same directory or accessible in the python path
from shared_utils import load_student_name_mapping

def load_room_no_mapping(filename):
    """
    Loads room name to room number mappings from a CSV file.
    The CSV should have 'room_name' and 'room_number' columns.
    """
    room_mappings = {}
    print(f"DEBUG - Loading room mapping from: {filename}")
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            print(f"DEBUG - CSV headers: {reader.fieldnames}")
            
            for row_num, row in enumerate(reader):
                room_name = row.get('room_name')
                room_number = row.get('room_number')
                print(f"DEBUG - Row {row_num + 1}: room_name='{room_name}', room_number='{room_number}'")
                
                if room_name and room_number:
                    room_mappings[room_name.strip()] = room_number.strip()
                    print(f"DEBUG - Added room mapping: '{room_name.strip()}' -> '{room_number.strip()}'")
                        
    except FileNotFoundError:
        print(f"Warning: Room mapping file '{filename}' not found.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    print(f"DEBUG - Final room mappings loaded: {room_mappings}")
    return room_mappings

def load_group_mapping(filename):
    """
    Loads group-to-student mappings from a CSV file.
    The CSV should have 'group_number' and 'student_no' columns.
    It creates a map where keys are group names (e.g., "Group 1") and
    values are lists of student IDs (e.g., ["F1", "C3", "H5"]).
    """
    group_mappings = {}
    print(f"DEBUG - Loading group mapping from: {filename}")
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            print(f"DEBUG - CSV headers: {reader.fieldnames}")
            
            for row_num, row in enumerate(reader):
                group_number = row.get('group_number')
                student_nos_str = row.get('student_no')
                print(f"DEBUG - Row {row_num + 1}: group_number='{group_number}', student_no='{student_nos_str}'")
                
                if group_number and student_nos_str:
                    group_name = f"Group {group_number.strip()}"
                    # Find all student IDs like F1, C12, H5
                    student_ids = re.findall(r'\b[FCH]\d+\b', student_nos_str)
                    print(f"DEBUG - Extracted student IDs: {student_ids}")
                    
                    if student_ids:
                        group_mappings[group_name] = student_ids
                        print(f"DEBUG - Added mapping: {group_name} -> {student_ids}")
                        
    except FileNotFoundError:
        print(f"Warning: Group mapping file '{filename}' not found.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    print(f"DEBUG - Final group mappings loaded: {group_mappings}")
    return group_mappings

def get_instrument_from_student_id(student_id):
    """Deduces the instrument name from the student ID prefix."""
    prefix = student_id[0].upper()
    if prefix == 'F':
        return 'Flute'
    elif prefix == 'C':
        return 'Cello'
    elif prefix == 'H':
        return 'Harp'
    return 'Unknown'

def update_master_timetable():
    """
    Main function to update the master timetable Excel file by replacing
    student and group numbers with detailed names and lists.
    """
    input_dir = "input"
    master_file_path = os.path.join(input_dir, "master.xlsx")
    output_file_path = os.path.join(input_dir, "master_processed.xlsx")
    
    # --- Target sheets to be processed ---
    target_sheets = [
        "Flute-Camp-A", "Cello-Camp-A", "Harp-Camp-A",
        "Flute-Camp-B", "Cello-Camp-B", "Harp-Camp-B"
    ]

    # --- Load all mapping files ---
    print("Loading mapping files...")
    student_map_a = load_student_name_mapping(os.path.join(input_dir, "student_mapping-campA.csv"))
    group_map_a = load_group_mapping(os.path.join(input_dir, "group_mapping-campA.csv"))
    room_no_map_a = load_room_no_mapping(os.path.join(input_dir, "room_no_mapping-campA.csv"))
    
    student_map_b = load_student_name_mapping(os.path.join(input_dir, "student_mapping-campB.csv"))
    group_map_b = load_group_mapping(os.path.join(input_dir, "group_mapping-campB.csv"))
    room_no_map_b = load_room_no_mapping(os.path.join(input_dir, "room_no_mapping-campB.csv"))

    # Debug: Print loaded mappings
    print(f"\nDEBUG - Camp A Student mappings loaded: {len(student_map_a)} entries")
    print(f"Sample student mappings: {dict(list(student_map_a.items())[:3])}")
    print(f"DEBUG - Camp A Group mappings loaded: {len(group_map_a)} entries")
    print(f"Sample group mappings: {dict(list(group_map_a.items())[:3])}")
    print(f"DEBUG - Camp A Room mappings loaded: {len(room_no_map_a)} entries")
    print(f"Sample room mappings: {dict(list(room_no_map_a.items())[:3])}")
    
    print(f"DEBUG - Camp B Student mappings loaded: {len(student_map_b)} entries")
    print(f"Sample student mappings: {dict(list(student_map_b.items())[:3])}")
    print(f"DEBUG - Camp B Group mappings loaded: {len(group_map_b)} entries")
    print(f"Sample group mappings: {dict(list(group_map_b.items())[:3])}")
    print(f"DEBUG - Camp B Room mappings loaded: {len(room_no_map_b)} entries")
    print(f"Sample room mappings: {dict(list(room_no_map_b.items())[:3])}")
    print()

    # --- Create a copy of the master file to preserve the original ---
    try:
        shutil.copy(master_file_path, output_file_path)
        print(f"Successfully copied '{master_file_path}' to '{output_file_path}'.")
    except FileNotFoundError:
        print(f"ERROR: Master file not found at '{master_file_path}'. Aborting.")
        return
    except Exception as e:
        print(f"ERROR: Could not copy file. {e}")
        return

    # --- Load the copied workbook for processing ---
    try:
        workbook = load_workbook(output_file_path)
    except Exception as e:
        print(f"ERROR: Could not load the workbook '{output_file_path}'. {e}")
        return

    print("Processing sheets...")
    print(f"DEBUG - Available sheets in workbook: {workbook.sheetnames}")
    print(f"DEBUG - Target sheets to process: {target_sheets}")
    
    for sheet_name in workbook.sheetnames:
        if sheet_name not in target_sheets:
            print(f"DEBUG - Skipping sheet '{sheet_name}' (not in target list)")
            continue

        print(f"\n  - Processing sheet: '{sheet_name}'")
        sheet = workbook[sheet_name]

        # Determine which set of mappings to use
        if "Camp-A" in sheet_name:
            student_map = student_map_a
            group_map = group_map_a
            room_no_map = room_no_map_a
        elif "Camp-B" in sheet_name:
            student_map = student_map_b
            group_map = group_map_b
            room_no_map = room_no_map_b
        else:
            continue # Should not happen with the current target_sheets list

        # Track cells that have been modified for resizing
        modified_cells = []
        cells_processed = 0
        cells_modified = 0

        # Iterate over every cell in the sheet
        for row in sheet.iter_rows():
            for cell in row:
                if not cell.value or not isinstance(cell.value, str):
                    continue
                
                cells_processed += 1
                cell_value = cell.value.strip()
                original_value = cell.value

                # Debug: Print every cell being processed (limit to first 10 for readability)
                if cells_processed <= 10:
                    print(f"DEBUG - Processing cell {cell.coordinate}: '{cell_value}'")

                # Priority 1: Check for an exact group match (e.g., "Group 1")
                group_match = re.fullmatch(r'Group\s+(\d+)', cell_value, re.IGNORECASE)
                if group_match:
                    print(f"DEBUG - Found exact group match in {cell.coordinate}: '{cell_value}'")
                    group_name = f"Group {group_match.group(1)}"
                    student_ids_in_group = group_map.get(group_name, [])
                    
                    print(f"DEBUG - Looking for group '{group_name}', found students: {student_ids_in_group}")
                    
                    if student_ids_in_group:
                        replacement_lines = [group_name]
                        for sid in student_ids_in_group:
                            student_name = student_map.get(sid, sid)
                            instrument = get_instrument_from_student_id(sid)
                            replacement_lines.append(f"{student_name}, {instrument}")
                            print(f"DEBUG - Replacing {sid} with {student_name}, {instrument}")
                        
                        cell.value = "\n".join(replacement_lines)
                        print(f"DEBUG - Cell {cell.coordinate} updated to: '{cell.value}'")

                        # Update alignment for readability, preserving as much as possible
                        new_alignment = copy(cell.alignment)
                        new_alignment.wrap_text = True
                        new_alignment.vertical = 'center'
                        new_alignment.horizontal = 'center'
                        cell.alignment = new_alignment
                        modified_cells.append(cell)
                        cells_modified += 1
                else:
                    # Priority 1.5: Check for complex group pattern (e.g., "Group 2, 5, 7, 9 Acting Class (Room Acting Class)")
                    complex_group_match = re.match(r'Group\s+(\d+(?:,\s*\d+)*)\s+(.+)', cell_value, re.IGNORECASE)
                    if complex_group_match:
                        print(f"DEBUG - Found complex group match in {cell.coordinate}: '{cell_value}'")
                        group_numbers_str = complex_group_match.group(1)
                        activity_with_room = complex_group_match.group(2).strip()
                        
                        print(f"DEBUG - Group numbers: '{group_numbers_str}', Activity: '{activity_with_room}'")
                        
                        # Extract individual group numbers
                        group_numbers = [num.strip() for num in group_numbers_str.split(',')]
                        
                        # Build the replacement
                        replacement_lines = []
                        
                        # Add activity and room info first
                        # Check if there's room info in parentheses
                        room_match = re.search(r'\(([^)]+)\)', activity_with_room)
                        if room_match:
                            activity_name = re.sub(r'\s*\([^)]+\)', '', activity_with_room).strip()
                            room_info = room_match.group(1)
                            replacement_lines.append(activity_name)
                            replacement_lines.append(f"({room_info})")
                        else:
                            replacement_lines.append(activity_with_room)
                        
                        # Add each group with its students
                        for group_num in group_numbers:
                            group_name = f"Group {group_num}"
                            student_ids_in_group = group_map.get(group_name, [])
                            
                            print(f"DEBUG - Looking for group '{group_name}', found students: {student_ids_in_group}")
                            
                            # Add group name on its own line
                            replacement_lines.append(group_name)
                            
                            if student_ids_in_group:
                                # Add student names on the next line
                                student_names = []
                                for sid in student_ids_in_group:
                                    student_name = student_map.get(sid, sid)
                                    student_names.append(student_name)
                                    print(f"DEBUG - Replacing {sid} with {student_name}")
                                replacement_lines.append(', '.join(student_names))
                        
                        cell.value = "\n".join(replacement_lines)
                        print(f"DEBUG - Cell {cell.coordinate} updated to: '{cell.value}'")

                        # Update alignment for readability
                        new_alignment = copy(cell.alignment)
                        new_alignment.wrap_text = True
                        new_alignment.vertical = 'center'
                        new_alignment.horizontal = 'center'
                        cell.alignment = new_alignment
                        modified_cells.append(cell)
                        cells_modified += 1
                    else:
                        # Priority 2: Replace student numbers within any other text
                        new_cell_value = cell.value
                        student_ids_found = re.findall(r'\b[FCH]\d+\b', new_cell_value)
                        
                        if student_ids_found:
                            print(f"DEBUG - Found student IDs in {cell.coordinate}: {student_ids_found}")
                        
                        # Use a set to avoid replacing the same ID multiple times if it appears more than once
                        for student_id in set(student_ids_found):
                            student_name = student_map.get(student_id, student_id)
                            print(f"DEBUG - Replacing student ID {student_id} with {student_name}")
                            # Use word boundary regex for safer replacement
                            new_cell_value = re.sub(r'\b' + re.escape(student_id) + r'\b', student_name, new_cell_value)
                        
                        if new_cell_value != cell.value:
                            cell.value = new_cell_value
                            print(f"DEBUG - Cell {cell.coordinate} updated from '{original_value}' to '{cell.value}'")
                            modified_cells.append(cell)
                            cells_modified += 1

                # Apply room mappings to all cells (final step after all other replacements)
                if cell.value and isinstance(cell.value, str):
                    original_cell_value = cell.value
                    updated_cell_value = cell.value
                    
                    # Replace room names with room numbers
                    for room_name, room_number in room_no_map.items():
                        if room_name in updated_cell_value:
                            updated_cell_value = updated_cell_value.replace(room_name, room_number)
                            print(f"DEBUG - Replacing room '{room_name}' with '{room_number}' in {cell.coordinate}")
                    
                    # Update cell if room mappings were applied
                    if updated_cell_value != original_cell_value:
                        cell.value = updated_cell_value
                        print(f"DEBUG - Cell {cell.coordinate} room mapping updated to: '{cell.value}'")
                        if cell not in modified_cells:
                            modified_cells.append(cell)
                            cells_modified += 1

        print(f"DEBUG - Sheet '{sheet_name}' summary: {cells_processed} cells processed, {cells_modified} cells modified")

        # Auto-resize columns to fit content
        print(f"    Auto-resizing columns for sheet '{sheet_name}'...")
        for column_cells in sheet.columns:
            max_length = 0
            # Handle merged cells - use the first non-merged cell to get column letter
            column_letter = None
            for cell in column_cells:
                if hasattr(cell, 'column_letter'):
                    column_letter = cell.column_letter
                    break
            
            # If we couldn't find a column letter, skip this column
            if not column_letter:
                continue
            
            for cell in column_cells:
                # Skip merged cells that don't have values
                if not hasattr(cell, 'value') or not cell.value:
                    continue
                    
                # Count lines and find the longest line
                lines = str(cell.value).split('\n')
                max_line_length = max(len(line) for line in lines) if lines else 0
                if max_line_length > max_length:
                    max_length = max_line_length
            
            # Set column width with padding (add 2 for padding)
            if max_length > 0:
                adjusted_width = min(max_length + 2, 100)  # Cap at 100 to prevent extremely wide columns
                sheet.column_dimensions[column_letter].width = adjusted_width

        # Auto-resize rows to fit content
        print(f"    Auto-resizing rows for sheet '{sheet_name}'...")
        for row in sheet.iter_rows():
            max_lines = 1
            row_number = None
            
            # Get row number from the first non-merged cell
            for cell in row:
                if hasattr(cell, 'row'):
                    row_number = cell.row
                    break
            
            # If we couldn't find a row number, skip this row
            if not row_number:
                continue
            
            for cell in row:
                # Skip merged cells that don't have values
                if not hasattr(cell, 'value') or not cell.value:
                    continue
                    
                # Count the number of lines in the cell
                lines = str(cell.value).count('\n') + 1
                if lines > max_lines:
                    max_lines = lines
            
            # Set row height based on number of lines (15 points per line + padding)
            if max_lines > 1:
                sheet.row_dimensions[row_number].height = max_lines * 18 + 5  # 18 points per line + 5 for padding
    
    # --- Save the updated workbook ---
    try:
        workbook.save(output_file_path)
        print(f"\nProcessing complete. Updated file saved to '{output_file_path}'.")
    except Exception as e:
        print(f"ERROR: Could not save the updated workbook. {e}")

if __name__ == '__main__':
    update_master_timetable() 
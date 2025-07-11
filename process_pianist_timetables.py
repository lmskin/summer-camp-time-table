import os
import re
import csv
import datetime
import itertools
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font, Border, Side
from openpyxl.utils import get_column_letter
from copy import copy

def load_student_name_mapping(filename):
    """
    Loads student ID to student name mappings from a CSV file.
    The CSV should have 'student_no' and 'student_name' columns.
    """
    student_mappings = {}
    print(f"DEBUG - Loading student mapping from: {filename}")
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            print(f"DEBUG - CSV headers: {reader.fieldnames}")
            
            for row_num, row in enumerate(reader):
                student_no = row.get('student_no')
                student_name = row.get('student_name')
                print(f"DEBUG - Row {row_num + 1}: student_no='{student_no}', student_name='{student_name}'")
                
                if student_no and student_name:
                    student_mappings[student_no.strip()] = student_name.strip()
                    print(f"DEBUG - Added student mapping: '{student_no.strip()}' -> '{student_name.strip()}'")
                        
    except FileNotFoundError:
        print(f"Warning: Student mapping file '{filename}' not found.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    print(f"DEBUG - Final student mappings loaded: {len(student_mappings)} entries")
    return student_mappings

def load_room_no_mapping(filename):
    """
    Loads room name to room number mappings from a CSV file.
    The CSV should have 'room_name' and 'room_number' columns.
    """
    room_mappings = {}
    print(f"DEBUG - Loading room mapping from: {filename}")
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            print(f"DEBUG - CSV headers: {reader.fieldnames}")
            
            for row_num, row in enumerate(reader):
                room_name = row.get('room_name')
                room_number = row.get('room_number')
                print(f"DEBUG - Row {row_num + 1}: room_name='{room_name}', room_number='{room_number}'")
                
                if room_name and room_number:
                    room_mappings[room_name.strip()] = room_number.strip()
                    print(f"DEBUG - Added room mapping: '{room_name.strip()}' -> '{room_number.strip()}'")
                        
    except FileNotFoundError:
        print(f"Warning: Room mapping file '{filename}' not found.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    print(f"DEBUG - Final room mappings loaded: {len(room_mappings)} entries")
    return room_mappings

def sanitize_filename(filename):
    """
    Remove or replace characters that are not valid in filenames.
    """
    # Replace problematic characters with underscores
    invalid_chars = '<>:"/\\|?*'
    for char in invalid_chars:
        filename = filename.replace(char, '_')
    return filename

def process_sheet_data(sheet, student_map, room_no_map):
    """
    Process sheet data into a structured format with proper time formatting.
    Returns: List of rows where each row is [time, col1_activity, col2_activity, ...]
    """
    processed_data = []
    
    # Get all data from the sheet
    max_row = sheet.max_row
    max_col = sheet.max_column
    
    # Process each row
    for row_num in range(1, max_row + 1):
        row_data = []
        
        for col_num in range(1, max_col + 1):
            cell = sheet.cell(row=row_num, column=col_num)
            value = cell.value
            
            # Handle time formatting for first column (assuming it's time)
            if col_num == 1 and value:
                if isinstance(value, datetime.time):
                    value = value.strftime('%H:%M')
                elif isinstance(value, datetime.datetime):
                    value = value.strftime('%H:%M')
                elif isinstance(value, str):
                    # Try to parse and reformat time strings
                    try:
                        # Handle various time formats
                        time_patterns = [
                            r'(\d{1,2}):(\d{2})',  # HH:MM or H:MM
                            r'(\d{1,2})\.(\d{2})',  # HH.MM or H.MM
                            r'(\d{1,2})(\d{2})',    # HHMM
                        ]
                        
                        time_str = str(value).strip()
                        parsed_time = None
                        
                        for pattern in time_patterns:
                            match = re.search(pattern, time_str)
                            if match:
                                hours = int(match.group(1))
                                minutes = int(match.group(2))
                                if 0 <= hours <= 23 and 0 <= minutes <= 59:
                                    parsed_time = f"{hours:02d}:{minutes:02d}"
                                    break
                        
                        if parsed_time:
                            value = parsed_time
                        
                    except (ValueError, AttributeError):
                        # Keep original value if parsing fails
                        pass
            
            # Process cell content for student ID and room replacements
            if value and isinstance(value, str):
                updated_value = value
                
                # Replace student IDs with student names
                student_ids_found = re.findall(r'\b[FCH]\d+\b', updated_value)
                for student_id in set(student_ids_found):
                    student_name = student_map.get(student_id, student_id)
                    if student_name != student_id:
                        updated_value = re.sub(r'\b' + re.escape(student_id) + r'\b', student_name, updated_value)
                
                # Replace room names with room numbers
                for room_name, room_number in room_no_map.items():
                    if room_name in updated_value:
                        updated_value = updated_value.replace(room_name, room_number)
                
                value = updated_value
            
            row_data.append(value)
        
        processed_data.append(row_data)
    
    return processed_data

def apply_cell_merging(pianist_ws, original_sheet):
    """
    Apply cell merging based on the original sheet's merged cell ranges
    and also merge consecutive empty cells in each column.
    This preserves the exact merged cell structure from the input file
    plus handles empty cell merging.
    Note: All ranges are shifted down by 2 rows to accommodate pianist name and date headers.
    """
    print("DEBUG - Starting cell merging process...")
    
    # Row shift offset to account for pianist name (row 1) and date headers (row 2)
    ROW_SHIFT = 2
    
    # First, copy original merged cell ranges (but skip row 1 to avoid conflicts with pianist name)
    if not original_sheet.merged_cells:
        print("DEBUG - No merged cells found in original sheet")
    else:
        print(f"DEBUG - Found {len(original_sheet.merged_cells.ranges)} merged cell ranges in original sheet")
        
        # Copy each merged cell range from the original sheet, adjusting for row shift
        for merged_range in original_sheet.merged_cells.ranges:
            try:
                # Get the range coordinates and shift them down by ROW_SHIFT
                min_col = merged_range.min_col
                max_col = merged_range.max_col
                min_row = merged_range.min_row + ROW_SHIFT
                max_row = merged_range.max_row + ROW_SHIFT
                
                # Skip merges that would conflict with pianist name (row 1) or date headers (row 2)
                if min_row <= 2:
                    print(f"DEBUG - Skipping merge {merged_range} because adjusted range {min_row}:{max_row} would conflict with headers")
                    continue
                
                # Only merge if it's actually a range (not a single cell)
                if min_row != max_row or min_col != max_col:
                    # Check if this conflicts with existing merges
                    conflict_found = False
                    for existing_range in pianist_ws.merged_cells.ranges:
                        if (not (max_row < existing_range.min_row or min_row > existing_range.max_row or
                                max_col < existing_range.min_col or min_col > existing_range.max_col)):
                            print(f"DEBUG - Merge conflict detected: adjusted {min_row}:{max_row},{min_col}:{max_col} conflicts with existing {existing_range}")
                            conflict_found = True
                            break
                    
                    if not conflict_found:
                        pianist_ws.merge_cells(start_row=min_row, start_column=min_col, 
                                             end_row=max_row, end_column=max_col)
                        print(f"DEBUG - Merged cells from original: {merged_range} -> adjusted to {min_row}:{max_row},{min_col}:{max_col}")
                    else:
                        print(f"DEBUG - Skipped conflicting merge: {merged_range}")
                
            except Exception as e:
                print(f"WARNING - Could not merge cells {merged_range}: {e}")
    
    # Apply evening time merging (19:00-22:00) for Monday to Friday (adjust for row shift)
    apply_evening_time_merging(pianist_ws, ROW_SHIFT)
    
    # Additional merging for consecutive empty cells in each column (skip rows 1-2)
    max_row = pianist_ws.max_row
    max_col = pianist_ws.max_column
    
    print("DEBUG - Applying additional merging for consecutive empty cells...")
    
    # Process each column for merging consecutive empty cells (skip first column which is time, and skip rows 1-2)
    for col_num in range(2, max_col + 1):
        print(f"DEBUG - Processing column {col_num} for empty cell merging...")
        
        # Create a list of (row_num, is_empty) pairs for this column (starting from row 3)
        column_data = []
        for row_num in range(3, max_row + 1):
            cell = pianist_ws.cell(row=row_num, column=col_num)
            is_empty = not cell.value or str(cell.value).strip() == "" or str(cell.value).strip().lower() == "none"
            column_data.append((row_num, is_empty))
        
        # Group consecutive empty cells and merge them
        for is_empty, group in itertools.groupby(column_data, key=lambda x: x[1]):
            if is_empty:  # Only merge empty cells
                group_list = list(group)
                if len(group_list) > 1:  # Only merge if there are multiple consecutive empty cells
                    start_row = group_list[0][0]
                    end_row = group_list[-1][0]
                    
                    try:
                        # Check if this range overlaps with any existing merged ranges
                        range_overlaps = False
                        for existing_range in pianist_ws.merged_cells.ranges:
                            if (existing_range.min_col <= col_num <= existing_range.max_col and
                                not (end_row < existing_range.min_row or start_row > existing_range.max_row)):
                                range_overlaps = True
                                print(f"DEBUG - Empty cell merge in column {col_num}, rows {start_row}-{end_row} overlaps with existing merge {existing_range}")
                                break
                        
                        if not range_overlaps:
                            # Merge the empty cells
                            pianist_ws.merge_cells(start_row=start_row, start_column=col_num, 
                                                 end_row=end_row, end_column=col_num)
                            print(f"DEBUG - Merged empty cells in column {col_num}, rows {start_row} to {end_row}")
                        else:
                            print(f"DEBUG - Skipped overlapping empty cell merge in column {col_num}, rows {start_row} to {end_row}")
                    except Exception as e:
                        print(f"WARNING - Could not merge empty cells in column {col_num}, rows {start_row} to {end_row}: {e}")

def apply_evening_time_merging(pianist_ws, row_shift=0):
    """
    Merge cells from 19:00 to 22:00 in the same column for Monday to Friday.
    This merges evening time slots regardless of content.
    """
    print("DEBUG - Applying evening time merging (19:00-22:00)...")
    
    # Define evening time slots to merge
    evening_times = ["19:00", "20:00", "21:00", "22:00"]
    
    # Find time column (column 1) to identify row numbers for evening times
    time_row_mapping = {}
    max_row = pianist_ws.max_row
    
    for row_num in range(1, max_row + 1):
        cell = pianist_ws.cell(row=row_num, column=1)
        if cell.value:
            time_str = str(cell.value).strip()
            # Handle various time formats
            if isinstance(cell.value, datetime.time):
                time_str = cell.value.strftime('%H:%M')
            elif ':' in time_str:
                # Already in HH:MM format
                pass
            elif '.' in time_str:
                time_str = time_str.replace('.', ':')
            elif len(time_str) == 4 and time_str.isdigit():
                # HHMM format
                time_str = f"{time_str[:2]}:{time_str[2:]}"
            elif len(time_str) == 3 and time_str.isdigit():
                # HMM format
                time_str = f"0{time_str[0]}:{time_str[1:]}"
            
            # Normalize to HH:MM format
            if ':' in time_str:
                parts = time_str.split(':')
                if len(parts) == 2:
                    try:
                        hour = int(parts[0])
                        minute = int(parts[1])
                        normalized_time = f"{hour:02d}:{minute:02d}"
                        time_row_mapping[normalized_time] = row_num
                    except ValueError:
                        pass
    
    # Find rows for evening times
    evening_rows = []
    for time_slot in evening_times:
        if time_slot in time_row_mapping:
            evening_rows.append(time_row_mapping[time_slot])
    
    if len(evening_rows) < 2:
        print(f"DEBUG - Not enough evening time slots found. Found: {evening_rows}")
        return
    
    # Sort evening rows to ensure correct order
    evening_rows.sort()
    
    print(f"DEBUG - Evening time rows to merge: {evening_rows} (times: {[time for time in evening_times if time in time_row_mapping]})")
    
    # Process each column (Monday to Friday) - columns 2 to 6 typically
    max_col = pianist_ws.max_column
    day_columns = range(2, min(max_col + 1, 7))  # Columns 2-6 for Mon-Fri
    
    for col_num in day_columns:
        column_letter = get_column_letter(col_num)
        
        # Check if this range overlaps with any existing merged ranges
        start_row = evening_rows[0]
        end_row = evening_rows[-1]
        
        range_overlaps = False
        for existing_range in pianist_ws.merged_cells.ranges:
            if (existing_range.min_col <= col_num <= existing_range.max_col and
                not (end_row < existing_range.min_row or start_row > existing_range.max_row)):
                range_overlaps = True
                print(f"DEBUG - Evening merge in column {col_num} ({column_letter}), rows {start_row}-{end_row} overlaps with existing merge {existing_range}")
                break
        
        if not range_overlaps:
            try:
                # Merge evening time slots in this column
                pianist_ws.merge_cells(start_row=start_row, start_column=col_num, 
                                     end_row=end_row, end_column=col_num)
                print(f"DEBUG - Merged evening times (19:00-22:00) in column {col_num} ({column_letter}), rows {start_row} to {end_row}")
            except Exception as e:
                print(f"WARNING - Could not merge evening times in column {col_num}: {e}")
        else:
            print(f"DEBUG - Skipped overlapping evening merge in column {col_num} ({column_letter}), rows {start_row} to {end_row}")

def process_pianist_timetables():
    """
    Main function to process pianist timetable Excel file by replacing
    student IDs with names and room names with numbers, then saving each sheet
    as a separate xlsx file with proper time formatting and cell merging.
    """
    pianist_dir = "Pianist Timetable"
    input_dir = "input"
    master_file_path = os.path.join(pianist_dir, "pianist-master-time-table.xlsx")
    output_dir = "pianist_timetables"
    
    # Create output directory if it doesn't exist
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)
        print(f"Created output directory: {output_dir}")

    # --- Load all mapping files ---
    print("Loading mapping files...")
    student_map_a = load_student_name_mapping(os.path.join(input_dir, "student_mapping-campA.csv"))
    room_no_map_a = load_room_no_mapping(os.path.join(input_dir, "room_no_mapping-campA.csv"))
    
    student_map_b = load_student_name_mapping(os.path.join(input_dir, "student_mapping-campB.csv"))
    room_no_map_b = load_room_no_mapping(os.path.join(input_dir, "room_no_mapping-campB.csv"))

    print(f"\nDEBUG - Camp A Student mappings loaded: {len(student_map_a)} entries")
    print(f"DEBUG - Camp A Room mappings loaded: {len(room_no_map_a)} entries")
    print(f"DEBUG - Camp B Student mappings loaded: {len(student_map_b)} entries")
    print(f"DEBUG - Camp B Room mappings loaded: {len(room_no_map_b)} entries")
    print()

    # --- Load the master workbook ---
    try:
        # Load without data_only to preserve merged cell information
        master_workbook = load_workbook(master_file_path, data_only=False)
        # Also load with data_only for actual cell values
        master_workbook_data = load_workbook(master_file_path, data_only=True)
        print(f"Successfully loaded pianist timetable from '{master_file_path}'")
    except FileNotFoundError:
        print(f"ERROR: Pianist timetable file not found at '{master_file_path}'. Aborting.")
        return
    except Exception as e:
        print(f"ERROR: Could not load the workbook '{master_file_path}'. {e}")
        return

    print("Processing sheets...")
    print(f"DEBUG - Available sheets in workbook: {master_workbook.sheetnames}")
    
    # Define border style
    thin_border = Border(
        left=Side(style='thin'),
        right=Side(style='thin'),
        top=Side(style='thin'),
        bottom=Side(style='thin')
    )

    for sheet_name in master_workbook.sheetnames:
        print(f"\n  - Processing sheet: '{sheet_name}'")
        
        # Check if sheet matches pianist name pattern (pianist name-campA/campB)
        camp_match = re.search(r"-(camp[ab])$", sheet_name, re.IGNORECASE)
        if not camp_match:
            print(f"DEBUG - Skipping sheet '{sheet_name}' (doesn't match pianist name-camp pattern)")
            continue

        # Extract pianist name and camp from sheet name
        parts = sheet_name.split('-')
        if len(parts) >= 2:  # Changed from >= 3 to >= 2 to handle names like 'Shelley_NG-campA'
            pianist_name = '-'.join(parts[:-1])  # All parts except the last (camp)
            camp_part = parts[-1].lower()
        else:
            print(f"WARNING: Sheet '{sheet_name}' does not follow expected naming convention")
            continue

        # Replace hyphens and underscores with spaces for display purposes
        pianist_display_name = pianist_name.replace('-', ' ').replace('_', ' ')
        
        print(f"DEBUG - Pianist: '{pianist_name}', Display name: '{pianist_display_name}', Camp: '{camp_part}'")

        # Determine which set of mappings to use and camp details
        if camp_part == "campa":
            student_map = student_map_a
            room_no_map = room_no_map_a
            camp_name = "CampA"
            start_date = datetime.date(2025, 7, 14)  # Camp A starts on 14 July 2025
        elif camp_part == "campb":
            student_map = student_map_b
            room_no_map = room_no_map_b
            camp_name = "CampB"
            start_date = datetime.date(2025, 7, 21)  # Camp B starts on 21 July 2025
        else:
            continue

        # Get both the original sheet (for structure) and data sheet (for values)
        original_sheet = master_workbook[sheet_name]  # For merged cell info
        data_sheet = master_workbook_data[sheet_name]  # For actual values
        
        # Process the data sheet for content
        processed_data = process_sheet_data(data_sheet, student_map, room_no_map)
        
        print(f"DEBUG - Processed {len(processed_data)} rows of data")

        # Create a new workbook for this pianist
        pianist_wb = Workbook()
        pianist_ws = pianist_wb.active
        pianist_ws.title = "Timetable"

        # Copy processed data to new worksheet, starting from row 3 to accommodate pianist name and date headers
        for row_num, row_data in enumerate(processed_data, start=3):
            for col_num, value in enumerate(row_data, start=1):
                cell = pianist_ws.cell(row=row_num, column=col_num)
                cell.value = value

        print(f"DEBUG - Copied {len(processed_data)} rows to new worksheet starting from row 3")

        # Clear row 1 completely and set up the pianist name
        max_col = pianist_ws.max_column
        print(f"DEBUG - Clearing row 1 and setting up pianist name across {max_col} columns")
        
        # Clear all cells in row 1 first
        for col in range(1, max_col + 1):
            pianist_ws.cell(row=1, column=col).value = None
        
        # Set pianist name only in the first cell
        first_cell = pianist_ws.cell(row=1, column=1)
        first_cell.value = pianist_display_name
        first_cell.font = Font(bold=True, size=20)
        
        # Merge pianist name across all columns in row 1 (only if there are multiple columns)
        if max_col > 1:
            try:
                pianist_ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=max_col)
                print(f"DEBUG - Successfully merged pianist name '{pianist_display_name}' across columns 1 to {max_col} in row 1")
            except Exception as e:
                print(f"WARNING - Could not merge pianist name across row 1: {e}")

        # Set up date headers in row 2 for columns 2-7 (Monday to Saturday)
        print(f"DEBUG - Setting up date headers for {camp_name} starting {start_date}")
        day_names = ["Monday", "Tuesday", "Wednesday", "Thursday", "Friday", "Saturday"]
        
        for day_index in range(6):  # Monday to Saturday (6 days)
            col_num = day_index + 2  # Columns 2-7
            if col_num <= max_col:
                current_date = start_date + datetime.timedelta(days=day_index)
                header_text = current_date.strftime('%d %B (%A)')  # Format: "14 July (Monday)"
                
                header_cell = pianist_ws.cell(row=2, column=col_num)
                header_cell.value = header_text
                header_cell.font = Font(bold=True, size=20)
                
                print(f"DEBUG - Set date header for column {col_num}: '{header_text}'")

        # Apply cell merging based on original sheet's merged cell structure and empty cells
        # Note: Need to adjust merge ranges since we shifted data down by 2 rows
        apply_cell_merging(pianist_ws, original_sheet)

        # Apply formatting: row heights and column widths
        print(f"    Applying formatting for sheet '{sheet_name}'...")
        
        # Set all row heights to 35
        for row_num in range(1, pianist_ws.max_row + 1):
            pianist_ws.row_dimensions[row_num].height = 35

        # Set column widths: time column auto-fit, date columns set to 80
        for col_num in range(1, pianist_ws.max_column + 1):
            column_letter = get_column_letter(col_num)
            
            if col_num == 1:  # Time column
                # Auto-fit time column based on content
                max_length = 0
                for row_num in range(1, pianist_ws.max_row + 1):
                    cell = pianist_ws.cell(row=row_num, column=col_num)
                    if cell.value:
                        lines = str(cell.value).split('\n')
                        max_line_length = max(len(line) for line in lines) if lines else 0
                        if max_line_length > max_length:
                            max_length = max_line_length
                
                # Set reasonable width for time column
                font_size_factor = 1.3
                padding = 2
                if max_length > 0:
                    adjusted_width = max(max_length * font_size_factor + padding, 15)
                    adjusted_width = min(adjusted_width, 25)  # Reasonable max for time column
                else:
                    adjusted_width = 15
                pianist_ws.column_dimensions[column_letter].width = adjusted_width
            else:  # Date columns
                # Set date columns to width 80
                pianist_ws.column_dimensions[column_letter].width = 80

        # Apply borders, alignment, and font to all cells
        for row in pianist_ws.iter_rows(min_row=1, max_row=pianist_ws.max_row, min_col=1, max_col=pianist_ws.max_column):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                # Apply 20pt font to all cells, preserving existing bold formatting if any
                if cell.font and cell.font.bold:
                    cell.font = Font(bold=True, size=20)
                else:
                    cell.font = Font(size=20)

        # Validate merged cells before saving
        print(f"DEBUG - Final merged cell count: {len(pianist_ws.merged_cells.ranges)}")
        for i, merged_range in enumerate(pianist_ws.merged_cells.ranges):
            print(f"DEBUG - Final merge {i+1}: {merged_range}")

        # Save the individual pianist timetable
        sanitized_pianist_name = sanitize_filename(pianist_name)
        output_file_path = os.path.join(output_dir, f"{sanitized_pianist_name}_{camp_name}_timetable.xlsx")
        
        try:
            pianist_wb.save(output_file_path)
            print(f"    Successfully saved '{output_file_path}'")
        except Exception as e:
            print(f"    ERROR: Could not save '{output_file_path}'. {e}")

    print(f"\nProcessing complete. Individual pianist timetables saved to '{output_dir}' directory.")

if __name__ == '__main__':
    process_pianist_timetables() 
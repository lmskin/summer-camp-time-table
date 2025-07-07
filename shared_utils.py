import re
import csv
from openpyxl import load_workbook
import os

def sanitize_filename(filename):
    """
    Removes characters from a string that are not allowed in file names.
    """
    return re.sub(r'[\\/*?:"<>|\n]', '', filename)

def process_sheet(sheet):
    """
    Processes a single sheet to extract its data, handling merged cells by
    unmerging them and filling the values.
    """
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_cell_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_cell_range.bounds
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(merged_cell_range))
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.value = top_left_cell_value

    data = []
    for row in sheet.iter_rows():
        data.append([cell.value if cell.value is not None else "" for cell in row])
    return data

def load_student_name_mapping(filename):
    """
    Loads student_no to student_name mappings from the specified CSV file.
    """
    print(f"DEBUG (shared_utils): Attempting to load student mapping from: {filename}")
    print(f"DEBUG (shared_utils): Current working directory: {os.getcwd()}")
    print(f"DEBUG (shared_utils): File exists check: {os.path.exists(filename)}")
    
    # List contents of input directory for debugging
    input_dir = os.path.dirname(filename)
    if os.path.exists(input_dir):
        print(f"DEBUG (shared_utils): Contents of {input_dir}:")
        for file in os.listdir(input_dir):
            print(f"  - {file}")
    else:
        print(f"DEBUG (shared_utils): Directory {input_dir} does not exist")
    
    name_map = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                student_no = row.get('student_no')
                student_name = row.get('student_name')
                if student_no and student_name:
                    name_map[student_no.strip()] = student_name.strip()
    except FileNotFoundError:
        print(f"Warning: {filename} not found. Student numbers will be used in filenames.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    print(f"DEBUG (shared_utils): Loaded student name map from {filename}. Found {len(name_map)} mappings.")
    if not name_map:
        print(f"DEBUG (shared_utils): The name map is empty. Check if the file exists and is correctly formatted.")

    return name_map

def load_room_no_mapping(filename):
    """
    Loads room_name to room_number mappings from the specified CSV file.
    """
    room_no_map = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                room_name = row.get('room_name')
                room_number = row.get('room_number')
                if room_name and room_number:
                    room_no_map[room_name.strip()] = room_number.strip()
    except FileNotFoundError:
        print(f"Warning: {filename} not found. Room names will not be replaced with numbers.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    return room_no_map 
"""
Script to adjust formatting of Excel files in the master_timetable folder.
Automatically adjusts column widths and row heights to fit cell content.
"""

import os
import glob
from openpyxl import load_workbook
from openpyxl.utils import get_column_letter
from openpyxl.styles import Alignment


def unmerge_day6_columns(worksheet):
    """
    Unmerge columns 2, 3, 4 in Day 6 sheet and move content to column 2.
    Preserve row merging for activities.
    
    Args:
        worksheet: The worksheet to process
    """
    print(f"    - DEBUG: *** FUNCTION CALLED - Starting unmerge process for Day 6 ***")
    
    # Get all merged cell ranges
    merged_ranges = list(worksheet.merged_cells.ranges)
    print(f"    - DEBUG: Found {len(merged_ranges)} merged ranges total")
    
    if len(merged_ranges) == 0:
        print(f"    - DEBUG: No merged ranges found - Day 6 sheet may have already been processed")
        return
    
    for i, merged_range in enumerate(merged_ranges):
        print(f"    - DEBUG: Range {i+1}: {merged_range} (rows {merged_range.min_row}-{merged_range.max_row}, cols {merged_range.min_col}-{merged_range.max_col})")
    
    ranges_to_remerge = []  # Store ranges that need to be re-merged in column 2
    processed_ranges = []
    
    for merged_range in merged_ranges:
        # Check if the merged range involves columns 2, 3, 4
        involves_target_columns = (merged_range.min_col <= 4 and merged_range.max_col >= 2 and 
                                 merged_range.min_col >= 2)
        
        print(f"    - DEBUG: Range {merged_range} involves columns 2-4: {involves_target_columns}")
        
        if involves_target_columns:
            # Get the content from the merged cell (top-left cell)
            top_left_cell = worksheet.cell(row=merged_range.min_row, column=merged_range.min_col)
            content = top_left_cell.value
            
            print(f"    - DEBUG: Content in range {merged_range}: '{content}'")
            print(f"    - DEBUG: Row span: {merged_range.min_row} to {merged_range.max_row} (spans {merged_range.max_row - merged_range.min_row + 1} rows)")
            
            # Store range info for re-merging in column 2 (if it spans multiple rows)
            if merged_range.min_row != merged_range.max_row:
                range_info = {
                    'min_row': merged_range.min_row,
                    'max_row': merged_range.max_row,
                    'content': content
                }
                ranges_to_remerge.append(range_info)
                print(f"    - DEBUG: Added to re-merge list: rows {range_info['min_row']}-{range_info['max_row']}")
            else:
                print(f"    - DEBUG: Single row range, will just move content")
            
            processed_ranges.append(merged_range)
    
    print(f"    - DEBUG: Will re-merge {len(ranges_to_remerge)} ranges in column 2")
    
    # Unmerge the processed ranges
    for merged_range in processed_ranges:
        print(f"    - DEBUG: Unmerging {merged_range}")
        worksheet.unmerge_cells(str(merged_range))
        
        # Clear all cells in the range
        for row in range(merged_range.min_row, merged_range.max_row + 1):
            for col in range(merged_range.min_col, merged_range.max_col + 1):
                cell = worksheet.cell(row=row, column=col)
                old_value = cell.value
                cell.value = None
                if old_value:
                    print(f"    - DEBUG: Cleared cell {get_column_letter(col)}{row} (was: '{old_value}')")
        
        print(f"    - DEBUG: Successfully unmerged {merged_range}")
    
    # Re-merge cells in column 2 for activities (preserving row spans)
    print(f"    - DEBUG: Starting re-merge process for {len(ranges_to_remerge)} ranges")
    for i, range_info in enumerate(ranges_to_remerge):
        print(f"    - DEBUG: Re-merge {i+1}: rows {range_info['min_row']}-{range_info['max_row']}")
        
        # Put the content in the top cell of column 2
        target_cell = worksheet.cell(row=range_info['min_row'], column=2)
        target_cell.value = range_info['content']
        print(f"    - DEBUG: Set content '{range_info['content']}' in B{range_info['min_row']}")
        
        # Re-merge the rows in column 2 only
        if range_info['min_row'] != range_info['max_row']:
            merge_range = f"B{range_info['min_row']}:B{range_info['max_row']}"
            print(f"    - DEBUG: Attempting to merge {merge_range}")
            try:
                worksheet.merge_cells(merge_range)
                print(f"    - SUCCESS: Re-merged rows in column 2: {merge_range}")
            except Exception as e:
                print(f"    - ERROR: Failed to merge {merge_range}: {str(e)}")
        else:
            print(f"    - DEBUG: Single row, just moved content to B{range_info['min_row']}")
    
    # Verify final state
    final_merged_ranges = list(worksheet.merged_cells.ranges)
    print(f"    - DEBUG: Final merged ranges count: {len(final_merged_ranges)}")
    for merged_range in final_merged_ranges:
        if merged_range.min_col == 2 and merged_range.max_col == 2:
            print(f"    - DEBUG: Column 2 merged range found: {merged_range}")
    
    print(f"    - DEBUG: *** FUNCTION COMPLETED - Unmerge process completed for Day 6 ***")


def adjust_column_width(worksheet, min_width=10, max_width=50):
    """
    Adjust column widths. Column 1 is auto-sized, all other columns are set to width 80.
    
    Args:
        worksheet: The worksheet to adjust
        min_width: Minimum column width for column 1
        max_width: Maximum column width for column 1
    """
    for column in worksheet.columns:
        column_letter = get_column_letter(column[0].column)
        column_number = column[0].column
        
        if column_number == 1:
            # Auto-size column 1 based on content
            max_length = 0
            for cell in column:
                try:
                    if cell.value:
                        # Calculate the length of the cell content
                        cell_length = len(str(cell.value))
                        if cell_length > max_length:
                            max_length = cell_length
                except:
                    pass
            
            # Set the column width with some padding
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            worksheet.column_dimensions[column_letter].width = adjusted_width
        else:
            # Set all other columns to width 80
            worksheet.column_dimensions[column_letter].width = 80


def adjust_row_height(worksheet, min_height=15, base_height=15):
    """
    Adjust row heights to ensure all content is fully visible.
    Calculate height based on content length, column width, and text wrapping.
    
    Args:
        worksheet: The worksheet to adjust
        min_height: Minimum row height
        base_height: Base height per line of text
    """
    for row_num in range(1, worksheet.max_row + 1):
        max_required_height = min_height
        
        for col_num in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row_num, column=col_num)
            if cell.value:
                content = str(cell.value)
                
                # Get column width
                column_letter = get_column_letter(col_num)
                if col_num == 1:
                    # Column 1 width varies, get actual width
                    col_width = worksheet.column_dimensions[column_letter].width or 10
                else:
                    # Other columns are set to 80
                    col_width = 80
                
                # Calculate approximate characters per line based on column width
                # Rough estimate: each character is about 1.2 units wide
                chars_per_line = max(1, int(col_width / 1.2))
                
                # Count explicit line breaks
                explicit_lines = content.count('\n') + 1
                
                # Calculate wrapped lines based on content length
                content_without_breaks = content.replace('\n', '')
                wrapped_lines = max(1, (len(content_without_breaks) + chars_per_line - 1) // chars_per_line)
                
                # Total lines is the sum of explicit breaks and wrapped content
                total_lines = explicit_lines + wrapped_lines - 1  # -1 because we counted base content twice
                
                # Calculate required height with vertical padding of 2
                required_height = max(total_lines * base_height + 2, min_height)
                
                if required_height > max_required_height:
                    max_required_height = required_height
        
        # Set the row height (no maximum limit to ensure content is never hidden)
        worksheet.row_dimensions[row_num].height = max_required_height


def apply_text_wrapping(worksheet):
    """
    Apply text wrapping and center alignment to all cells in the worksheet.
    This ensures consistent formatting and proper height calculations.
    
    Args:
        worksheet: The worksheet to apply wrapping and alignment to
    """
    # Apply formatting to all cells in the used range
    for row in range(1, worksheet.max_row + 1):
        for col in range(1, worksheet.max_column + 1):
            cell = worksheet.cell(row=row, column=col)
            # Apply alignment to all cells (with or without content)
            cell.alignment = Alignment(wrap_text=True, vertical='center', horizontal='center')


def format_excel_file(file_path):
    """
    Format a single Excel file by adjusting column widths and row heights.
    
    Args:
        file_path: Path to the Excel file to format
    """
    try:
        print(f"Processing: {os.path.basename(file_path)}")
        
        # Load the workbook
        workbook = load_workbook(file_path)
        
        # Process each worksheet
        for sheet_name in workbook.sheetnames:
            print(f"  - Formatting sheet: {sheet_name}")
            worksheet = workbook[sheet_name]
            
            # Special handling for Day 6 sheet
            if "Day 6" in sheet_name or sheet_name == "Day 6":
                print(f"    - Applying special Day 6 formatting")
                unmerge_day6_columns(worksheet)
            
            # Apply formatting in optimal order
            # 1. First apply text wrapping and alignment to all cells
            apply_text_wrapping(worksheet)
            # 2. Set column widths
            adjust_column_width(worksheet)
            # 3. Finally adjust row heights based on final column widths
            adjust_row_height(worksheet)
        
        # Create backup before saving
        backup_path = file_path.replace('.xlsx', '_formatted.xlsx')
        if not os.path.exists(backup_path):
            workbook.save(backup_path)
            print(f"  - Backup created: {os.path.basename(backup_path)}")
        
        # Save the formatted file
        workbook.save(file_path)
        print(f"  - Formatting completed for: {os.path.basename(file_path)}")
        
    except Exception as e:
        print(f"Error processing {file_path}: {str(e)}")


def format_master_timetables(folder_path):
    """
    Format all Excel files in the master_timetable folder.
    
    Args:
        folder_path: Path to the master_timetable folder
    """
    if not os.path.exists(folder_path):
        print(f"Error: Folder '{folder_path}' does not exist.")
        return
    
    # Find all Excel files in the folder
    excel_files = glob.glob(os.path.join(folder_path, "*.xlsx"))
    
    # Filter out temporary files (starting with ~$)
    excel_files = [f for f in excel_files if not os.path.basename(f).startswith('~$')]
    
    if not excel_files:
        print(f"No Excel files found in '{folder_path}'")
        return
    
    print(f"Found {len(excel_files)} Excel file(s) to format:")
    for file_path in excel_files:
        print(f"  - {os.path.basename(file_path)}")
    
    print("\nStarting formatting process...")
    print("=" * 50)
    
    # Process each file
    for file_path in excel_files:
        format_excel_file(file_path)
        print("-" * 30)
    
    print("Formatting process completed!")


def main():
    """Main function to run the formatting script."""
    try:
        # Get the current script directory
        script_dir = os.path.dirname(os.path.abspath(__file__))
        
        # Define the master_timetable folder path
        master_timetable_folder = os.path.join(script_dir, "master_timetable")
        
        print("Excel File Formatter for Master Timetables")
        print("=" * 50)
        print(f"Script directory: {script_dir}")
        print(f"Target folder: {master_timetable_folder}")
        print(f"Folder exists: {os.path.exists(master_timetable_folder)}")
        print()
        
        # Format the files
        format_master_timetables(master_timetable_folder)
        
    except Exception as e:
        print(f"Error in main function: {str(e)}")
        import traceback
        traceback.print_exc()


if __name__ == "__main__":
    main()

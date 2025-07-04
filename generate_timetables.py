import re
import os
import itertools
import datetime
import csv
from openpyxl import load_workbook, Workbook
from openpyxl.styles import Alignment, Font
from openpyxl.utils import get_column_letter

def sanitize_filename(filename):
    """
    Removes characters from a string that are not allowed in file names.
    This includes newline characters, which can cause issues with file paths.
    """
    return re.sub(r'[\\/*?:"<>|\\n]', '', filename)

def process_sheet(sheet):
    """
    Processes a single sheet to extract its data, handling merged cells by
    unmerging them and filling the values.
    """
    # Create a copy of the merged cell ranges to iterate over, as unmerging
    # will modify the sheet's merged_cells attribute.
    merged_ranges = list(sheet.merged_cells.ranges)
    for merged_cell_range in merged_ranges:
        min_col, min_row, max_col, max_row = merged_cell_range.bounds
        top_left_cell_value = sheet.cell(row=min_row, column=min_col).value
        sheet.unmerge_cells(str(merged_cell_range))
        for row in sheet.iter_rows(min_row=min_row, max_row=max_row, min_col=min_col, max_col=max_col):
            for cell in row:
                cell.value = top_left_cell_value

    # Read the data from the now unmerged sheet into a 2D list
    data = []
    for row in sheet.iter_rows():
        data.append([cell.value if cell.value is not None else "" for cell in row])
    return data

def load_student_name_mapping(filename="student_mapping.csv"):
    """
    Loads student_no to student_name mappings from the specified CSV file.
    """
    name_map = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                student_no = row.get('student_no')
                student_name = row.get('student_name')
                if student_no and student_name:
                    name_map[student_no.strip()] = student_name.strip()
    except FileNotFoundError:
        print(f"Warning: {filename} not found. Student numbers will be used in filenames.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    return name_map

def load_group_mappings(filename="group_mapping.csv"):
    """
    Loads group-to-student mappings from the specified CSV file.
    The CSV should have 'group_number' and 'student_no' columns.
    It extracts student IDs (e.g., F1) from the 'student_no' column.
    """
    group_mappings = {}
    try:
        with open(filename, mode='r', encoding='utf-8-sig') as infile:
            reader = csv.DictReader(infile)
            for row in reader:
                group_number = row.get('group_number')
                student_nos_str = row.get('student_no')

                if group_number and student_nos_str:
                    group_name = f"Group {group_number.strip()}"
                    # Extract all F-numbers (e.g., F1, F23) from the string
                    found_students = re.findall(r'\bF\d+\b', student_nos_str)
                    
                    if found_students:
                        if group_name not in group_mappings:
                            group_mappings[group_name] = []
                        group_mappings[group_name].extend(found_students)
    except FileNotFoundError:
        print(f"Warning: {filename} not found. No group activities will be added.")
    except Exception as e:
        print(f"An error occurred while reading {filename}: {e}")
    
    return group_mappings

def generate_timetables():
    """
    Reads an Excel file with multiple sheets (each representing a date) and
    generates an individual Excel timetable for each student. Each student's
    timetable will contain multiple sheets, corresponding to the dates in the
    input file.
    """
    input_filename = "flute-time-table.xlsx"
    
    # Extract instrument name, e.g., 'cello' from 'cello-time-table.xlsx'
    music_instrument = input_filename.replace('-time-table.xlsx', '')
    
    try:
        # Load the workbook once to process all sheets.
        # data_only=True ensures we get cell values instead of formulas.
        workbook = load_workbook(input_filename, data_only=True)
        print(f"Processing timetable for {music_instrument.capitalize()}...")
    except FileNotFoundError:
        # This case is less likely now but good to keep as a safeguard
        print(f"Error: {input_filename} not found. Make sure the file is in the same directory.")
        return

    # Load mappings
    student_name_map = load_student_name_mapping()
    group_mappings = load_group_mappings()
    
    # Create a reverse mapping from student to their groups for efficient lookup
    student_to_groups = {}
    for group, students in group_mappings.items():
        for s in students:
            if s not in student_to_groups:
                student_to_groups[s] = set()
            student_to_groups[s].add(group)

    # Process all sheets and store their data in a dictionary
    processed_sheets = {}
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]
        processed_sheets[sheet_name] = process_sheet(sheet)

    # Find all unique students across all processed sheets
    all_students = set()
    for sheet_name, sheet_data in processed_sheets.items():
        schedule_rows = sheet_data[2:]  # Schedule data starts from the third row
        for row in schedule_rows:
            for cell in row[1:]:
                if isinstance(cell, str):
                    # Use regex to find all student IDs (e.g., F1) in a cell
                    found_students = re.findall(r'\bF\d+\b', cell)
                    for s in found_students:
                        all_students.add(s)

    common_activities = [
        "Welcome Speech",
        "Lunch",
        "Break",
        "Ensemble Coaching",
        "Workshop",
        "Toilet Break",
        "Rehearsal for Students and Friends Concert",
        "Lina Summer Camp of Music Students & Friends Concert",
        "After concert refreshment (Maritime Museum)",
    ]

    output_dir = "student_timetables"
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Generate one single-sheet Excel file for each student
    for student in sorted(list(all_students)):
        # Pre-process to gather all time slots and daily schedules for the student
        all_time_slots = set()
        daily_schedules = {}

        # Using workbook.sheetnames to preserve the order of days
        for day_index, sheet_name in enumerate(workbook.sheetnames):
            is_day_6 = (day_index + 1 == 6)

            if sheet_name not in processed_sheets:
                continue
            
            sheet_data = processed_sheets[sheet_name]
            teachers = [str(name).strip() for name in sheet_data[0][1:]]
            schedule_rows = sheet_data[2:]
            
            student_schedule = []
            for row in schedule_rows:
                time_val = row[0]
                time = time_val.strftime('%H:%M') if isinstance(time_val, datetime.time) else str(time_val).strip()

                # For Day 1 to Day 5, students finish at 17:00, so no activities are scheduled from 17:00 onwards.
                if day_index < 5:  # Day 1 to 5
                    try:
                        # We only compare times that are in the HH:MM format
                        if datetime.datetime.strptime(time, '%H:%M').time() >= datetime.time(17, 0):
                            continue  # Skip this timeslot
                    except ValueError:
                        pass  # Not a time in HH:MM format, so we don't apply the filter

                activities = [str(act).strip() for act in row[1:]]
                activity_found_for_timeslot = False

                if is_day_6:
                    # For Day 6, any activity is considered a common activity for all students.
                    # Find the first non-empty activity in the row.
                    for activity in activities:
                        if activity:
                            student_schedule.append((time, activity))
                            activity_found_for_timeslot = True
                            break # Move to the next timeslot
                else:
                    # For all other days, run the specific matching logic.
                    for i, activity in enumerate(activities):
                        if not activity:
                            continue

                        # Priority 1: Direct student match
                        if not activity_found_for_timeslot and re.search(r'\b' + re.escape(student) + r'\b', activity):
                            if activity.strip() == student:
                                teacher = teachers[i]
                                desc = f"Private lesson with {teacher}" if teacher else f"Practice ({music_instrument} practice room)"
                                student_schedule.append((time, desc))
                            else:
                                # Remove student's own ID from the activity description, as the timetable is already individualized.
                                cleaned_activity = re.sub(r'\b' + re.escape(student) + r'\b', '', activity).strip()
                                if cleaned_activity.lower() == 'practice':
                                    cleaned_activity = f"Practice ({music_instrument} practice room)"
                                student_schedule.append((time, cleaned_activity))
                            activity_found_for_timeslot = True

                        # Priority 2: Complex group match
                        if not activity_found_for_timeslot and activity.lower().startswith('group') and "," in activity:
                            activity_body = activity[len('Group'):].strip()
                            
                            parts = activity_body.replace(',', ' ').split()
                            
                            group_numbers = []
                            activity_name_parts = []
                            for part in parts:
                                if part.isdigit():
                                    group_numbers.append(part)
                                else:
                                    activity_name_parts.append(part)
                            
                            activity_name = ' '.join(activity_name_parts).strip()
                            involved_groups = {f"Group {num}" for num in group_numbers}

                            student_groups = student_to_groups.get(student, set())

                            if not student_groups.isdisjoint(involved_groups):
                                if 'acting class' in activity_name.lower():
                                    student_schedule.append((time, "Acting class"))
                                else:
                                    student_schedule.append((time, f"{activity_name} (Group)"))
                                activity_found_for_timeslot = True

                        # Priority 3: Simple group match (e.g., "Group 1")
                        if not activity_found_for_timeslot and activity.lower().startswith('group'):
                            student_groups = student_to_groups.get(student, set())
                            if activity in student_groups:
                                student_schedule.append((time, "Ensemble"))
                                activity_found_for_timeslot = True
                
                # Fallback for common activities or Free Time
                if not activity_found_for_timeslot:
                    # Check for any common activity for this timeslot.
                    # Master Class is checked first, then the predefined list.
                    activity_to_add = None
                    for activity in activities:
                        if activity.startswith("Master class with"):
                            activity_to_add = activity
                            break  # Found master class, stop searching this row
                    
                    if not activity_to_add:
                        for common_activity in common_activities:
                            if common_activity in activities:
                                activity_to_add = common_activity
                                break  # Found a common activity
    
                    if activity_to_add:
                        student_schedule.append((time, activity_to_add))
                    else:
                        student_schedule.append((time, "Free Time"))

            daily_schedules[sheet_name] = student_schedule
            for time, _ in student_schedule:
                all_time_slots.add(time)
        
        sorted_times = sorted(list(all_time_slots))
        time_to_row = {time: i + 3 for i, time in enumerate(sorted_times)}

        student_wb = Workbook()
        student_ws = student_wb.active
        student_ws.title = "Full Timetable"

        student_ws.cell(row=2, column=1, value="Time").font = Font(bold=True)
        for i, time in enumerate(sorted_times):
            student_ws.cell(row=i + 3, column=1, value=time)

        current_col = 2
        for sheet_name in workbook.sheetnames:
            if sheet_name not in daily_schedules:
                continue
            
            student_ws.cell(row=1, column=current_col, value=sheet_name).font = Font(bold=True)

            todays_schedule = daily_schedules[sheet_name]

            for activity, group in itertools.groupby(todays_schedule, key=lambda x: x[1]):
                group_list = list(group)
                row_span = len(group_list)
                start_time = group_list[0][0]
                
                if start_time not in time_to_row: continue

                start_row = time_to_row[start_time]
                cell = student_ws.cell(row=start_row, column=current_col, value=activity)

                if row_span > 1:
                    end_row = start_row + row_span - 1
                    student_ws.merge_cells(start_row=start_row, start_column=current_col, end_row=end_row, end_column=current_col)
                    cell.alignment = Alignment(vertical='center')

            current_col += 2

        # Autofit all columns
        for column in student_ws.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = (max_length + 2)
            student_ws.column_dimensions[column_letter].width = adjusted_width

        # Use student name for the filename, falling back to student number
        student_name = student_name_map.get(student, student)
        sanitized_file_name = sanitize_filename(student_name)
        file_path = os.path.join(output_dir, f'{sanitized_file_name}_timetable.xlsx')
        student_wb.save(file_path)

    print(f"Successfully generated timetables for {len(all_students)} students in the '{output_dir}' directory.")


if __name__ == '__main__':
    generate_timetables() 